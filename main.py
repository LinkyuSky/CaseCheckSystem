# ==============================================
# 检察机关案卡冲突自动审核系统
# 架构：模块化规则引擎 - 新增规则只加函数，不动主程序
# ==============================================

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime
import os
import sys
import shutil

BUILD_DATE = "2026-04-29"

def check_system_date():
    try:
        build_date = datetime.strptime(BUILD_DATE, "%Y-%m-%d")
        current_date = datetime.now()
        
        if current_date.date() < build_date.date():
            import tkinter as tk
            from tkinter import ttk
            
            dialog = tk.Tk()
            dialog.title("系统日期异常")
            dialog.geometry("600x360")
            dialog.resizable(False, False)
            dialog.configure(bg="#f8fafc")
            
            dialog.overrideredirect(False)
            dialog.attributes("-topmost", True)
            
            header = tk.Frame(dialog, bg="#1e3a8a", height=80)
            header.pack(fill="x")
            header.pack_propagate(False)
            
            icon_label = tk.Label(header, text="⚠️", font=("微软雅黑", 32), bg="#1e3a8a", fg="white")
            icon_label.pack(side="left", padx=20, pady=15)
            
            title_label = tk.Label(header, text="系统日期异常", font=("微软雅黑", 18, "bold"), bg="#1e3a8a", fg="white")
            title_label.pack(side="left", pady=25)
            
            content = tk.Frame(dialog, bg="#f8fafc", padx=30, pady=25)
            content.pack(fill="both", expand=True)
            
            msg_text = f"""检测到系统日期异常，请修改系统日期

当前系统日期: {current_date.strftime('%Y-%m-%d')}
"""
            msg_label = tk.Label(content, text=msg_text, font=("微软雅黑", 16, "bold"), bg="#f8fafc", fg="black", justify="left", padx=40, pady=0)
            msg_label.pack(fill="x", pady=(20, 0))
            
            btn_frame = tk.Frame(dialog, bg="#1e3a8a")
            btn_frame.pack(fill="x", padx=(360, 60), pady=(0, 40))
            
            def on_close():
                dialog.destroy()
                sys.exit(1)
            
            close_btn = tk.Button(btn_frame, text="确定", command=on_close,
                                  font=("微软雅黑", 12, "bold"), bg="#2563eb", fg="white",
                                  bd=0, padx=40, pady=10,
                                  activebackground="#1d4ed8", cursor="hand2")
            close_btn.pack(fill="x")
            
            close_btn.bind("<Enter>", lambda e: close_btn.config(bg="#1d4ed8"))
            close_btn.bind("<Leave>", lambda e: close_btn.config(bg="#2563eb"))
            
            dialog.protocol("WM_DELETE_WINDOW", on_close)
            dialog.mainloop()
            
            sys.exit(1)
    except Exception as e:
        pass

check_system_date()

# ====================== 授权模块 ======================
from auth_new import check_license, activate_new, generate_short_identifier, is_license_valid

# ====================== 窗口配置 ======================
root = tk.Tk()
root.title("检察机关案卡冲突排查系统")
root.geometry("1450x950")
root.resizable(True, True)
root.minsize(1200, 850)

# ====================== 文件配置 ======================
FILE_ARREST = "审查逮捕案件010101表1行4列受理合计（人）.xlsx"
FILE_PROSEC = "一审公诉案件020101表1行19列受理合计（人）.xlsx"
FILE_FILTER = "已核对清单.xlsx"

path_arrest = ""
path_prosec = ""
path_filter = ""

# ====================== 界面控件 ======================
label_arrest = None
label_prosec = None
label_filter = None
label_arrest_count = None
label_prosec_count = None
label_filter_count = None
label_arrest_tip = None
label_prosec_tip = None
label_filter_tip = None
label_auth_info = None
status_vars = {}
status_label = None

# ====================== 自定义对话框 ======================
def show_custom_dialog(title, message, icon_type="warning"):
    dialog = tk.Toplevel(root)
    dialog.title(title)
    dialog.geometry("520x220")
    dialog.resizable(False, False)
    dialog.transient(root)
    dialog.grab_set()
    
    bg_color = "#ffffff"
    dialog.configure(bg="#f0f4f8")
    
    icon_text = ""
    accent_color = "#f59e0b"
    if icon_type == "warning":
        icon_text = "⚠️"
        accent_color = "#f59e0b"
    elif icon_type == "error":
        icon_text = "❌"
        accent_color = "#ef4444"
    elif icon_type == "info":
        icon_text = "ℹ️"
        accent_color = "#3b82f6"
    elif icon_type == "success":
        icon_text = "✅"
        accent_color = "#22c55e"
    
    content = tk.Frame(dialog, bg=bg_color, padx=30, pady=25)
    content.pack(fill="both", expand=True)
    
    header = tk.Frame(content, bg=bg_color)
    header.pack(fill="x", pady=(0, 10))
    
    icon_label = tk.Label(header, text=icon_text, font=("Microsoft YaHei", 36), bg=bg_color)
    icon_label.pack(side="left", padx=(0, 20))
    
    title_label = tk.Label(header, text=title, font=("Microsoft YaHei", 16, "bold"), bg=bg_color, fg="#1f2937")
    title_label.pack(side="left", anchor="w")
    
    msg_label = tk.Label(content, text=message, font=("Microsoft YaHei", 11), bg=bg_color, fg="#4b5563", wraplength=420)
    msg_label.pack(fill="x", pady=8, anchor="w")
    
    btn_frame = tk.Frame(content, bg=bg_color)
    btn_frame.pack(fill="x", pady=(10, 0), side="bottom")
    
    def on_ok():
        dialog.destroy()
    
    ok_btn = tk.Button(btn_frame, text="确定", font=("微软雅黑", 12, "bold"), bg=accent_color, fg="white",
                      activebackground=accent_color, activeforeground="white",
                      padx=45, pady=20, relief="flat", command=on_ok, cursor="hand2")
    ok_btn.pack(side="right")
    
    center_window(dialog)
    root.wait_window(dialog)


def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f"+{x}+{y}")


# ====================== 通用字段 ======================
COMMON_FIELDS = ["部门受案号", "案件名称", "嫌疑人姓名", "受理日期", "承办部门", "承办检察官", "嫌疑人编号"]

# ====================== 公用工具体 ======================
def clean(s):
    return str(s).strip() if pd.notna(s) else ""

def parse_charges(s):
    return set(x.strip() for x in clean(s).split(",") if x.strip())

def build_error_result(row, rule_code, rule_name, conflict_msg, source_table="010101表", **extra_fields):
    result = {k: clean(row.get(k)) for k in COMMON_FIELDS}
    result.update(extra_fields)
    result["规则编号"] = rule_code
    result["规则名称"] = rule_name
    result["冲突说明"] = conflict_msg
    result["来源表"] = source_table
    return result

# ====================== 规则常量 ======================
OFFICIAL_CRIMES = [
    "贪污罪", "挪用公款罪", "受贿罪", "单位受贿罪", "行贿罪", "对单位行贿罪", "介绍贿赂罪","巨额财产来源不明罪", "隐瞒境外存款罪", "私分国有资产罪", 
    "私分罚没财物罪", "利用影响力受贿罪","对有影响力的人行贿罪", "滥用职权罪", "玩忽职守罪", "故意泄露国家秘密罪", "过失泄露国家秘密罪","徇私枉法罪", 
    "民事、行政枉法裁判罪", "私放在押人员罪", "失职致使在押人员脱逃罪","徇私舞弊减刑、假释、暂予监外执行罪", "徇私舞弊不移交刑事案件罪", "滥用管理公司、证券职权罪",
    "徇私舞弊不征、少征税款罪", "徇私舞弊发售发票、抵扣税款、出口退税罪", "环境监管失职罪","传染病防治失职罪", "国家机关工作人员签订、履行合同失职被骗罪", 
    "违法发放林木采伐许可证罪","非法批准征收、征用、占用土地罪（原非法批准征用、占用土地罪）", "放纵走私罪", "商检徇私舞弊罪","商检失职罪", "动植物检疫徇私舞弊罪", 
    "动植物检疫失职罪", "放纵制售伪劣商品犯罪行为罪","办理偷越国(边)境人员出入境证件罪", "放行偷越国(边)境人员罪", "不解救被拐卖、绑架妇女、儿童罪",
    "阻碍解救被拐卖、绑架妇女、儿童罪", "帮助犯罪分子逃避处罚罪", "招收公务员、学生徇私舞弊罪", "失职造成珍贵文物损毁、流失罪", "违法提供出口退税凭证罪", 
    "非法低价出让国有土地使用权罪","枉法仲裁罪", "食品、药品监管渎职罪（原食品监管渎职罪）", "执行判决、裁定失职罪","执行判决、裁定滥用职权罪", "国家工作人员非法拘禁", 
    "国家工作人员非法搜查", "国家工作人员刑讯逼供","暴力取证", "报复陷害", "虐待被监管人", "破坏选举", "其他利用职权犯罪", "战时违抗命令罪","隐瞒、谎报军情罪", 
    "拒传、假传军令罪", "投降罪", "战时临阵脱逃罪", "擅离、玩忽军事职守罪","阻碍执行军事职务罪", "指使部属违反职责罪", "违令作战消极罪", "拒不救援友邻部队罪",
    "军人叛逃罪", "非法获取军事秘密罪", "为境外窃取、刺探、收买、非法提供军事秘密罪","故意泄露军事秘密罪", "过失泄露军事秘密罪", "战时造谣惑众罪", "战时自伤罪", 
    "逃离部队罪","武器装备肇事罪", "擅自改变武器装备编配用途罪", "盗窃、抢夺武器装备、军用物资罪","非法出卖、转让武器装备罪", "遗弃武器装备罪", "遗失武器装备罪", 
    "擅自出卖、转让军队房地产罪","虐待部属罪", "遗弃伤病军人罪", "战时拒不救治伤病军人罪", "战时残害居民、掠夺居民财物罪","私放俘虏罪", "虐待俘虏罪",
]

SPECIAL_CRIMES = [
    "背叛国家罪", "分裂国家罪", "煽动分裂国家罪", "武装叛乱、暴乱罪", "颠覆国家政权罪","煽动颠覆国家政权罪", "资助危害国家安全犯罪活动罪", "投敌叛变罪", "叛逃罪", 
    "间谍罪","资敌罪", "组织、领导、参加恐怖组织罪", "帮助恐怖活动罪", "准备实施恐怖活动罪","宣扬恐怖主义、极端主义、煽动实施恐怖活动罪",
    "强制穿戴宣扬恐怖主义、极端主义服饰、标志罪","非法持有宣扬恐怖主义、极端主义物品罪", "编造、故意传播虚假恐怖信息罪",
    "拒绝提供间谍犯罪、恐怖主义犯罪、极端主义犯罪证据罪","为境外窃取、刺探、收买、非法提供国家秘密、情报罪",
]

# ====================== 经济类、侵财类犯罪 =====================
ECONOMIC_CRIMES = [
    "生产、销售伪劣商品罪", "生产、销售伪劣产品罪", "生产、销售不符合安全标准的食品罪","生产、销售有毒、有害食品罪", "生产、销售不符合标准的医用器材罪",
    "生产、销售不符合安全标准的产品罪", "生产、销售伪劣农药、兽药、化肥、种子罪","生产、销售不符合卫生标准的化妆品罪", "妨害药品管理罪", "走私罪",
    "走私武器、弹药罪", "走私核材料罪", "走私假币罪", "走私文物罪", "走私贵重金属罪","走私珍贵动物、珍贵动物制品罪", "走私国家禁止进出口的货物、物品罪",
    "走私淫秽物品罪", "走私普通货物、物品罪", "走私废物罪", "妨害对公司、企业的管理秩序罪","虚报注册资本罪", "虚假出资、抽逃出资罪", 
    "欺诈发行证券罪（原欺诈发行股票、债券罪）","违规披露、不披露重要信息罪", "隐匿、故意销毁会计凭证、会计帐簿、财务会计报告罪","妨害清算罪", 
    "非国家工作人员受贿罪", "对非国家工作人员行贿罪", "非法经营同类营业罪","为亲友非法牟利罪", "签订、履行合同失职被骗罪", "国有公司、企业、事业单位人员失职罪",
    "国有公司、企业、事业单位人员滥用职权罪", "徇私舞弊低价折股、出售国有资产罪","虚假破产罪", "背信损害上市公司利益罪", "对外国公职人员、国际公共组织官员行贿罪",
    "破坏金融管理秩序罪", "伪造货币罪", "出售、购买、运输假币罪","金融工作人员购买假币、以假币换取货币罪", "持有、使用假币罪","变造货币罪", "擅自设立金融机构罪", 
    "高利转贷罪", "非法吸收公众存款罪","伪造、变造金融票证罪", "伪造、变造国家有价证券罪", "伪造、变造股票、公司、企业债券罪","擅自发行股票、公司、企业债券罪", 
    "内幕交易、泄露内幕信息罪","编造并传播证券、期货交易虚假信息罪", "诱骗投资者买卖证券、期货合约罪","操纵证券、期货市场罪", "违法发放贷款罪", 
    "吸收客户资金不入账罪","违规出具金融票证罪", "对违法票据承兑、付款、保证罪", "逃汇罪", "洗钱罪","骗购外汇罪", "骗取贷款、票据承兑、金融票证罪", 
    "妨害信用卡管理罪","窃取、收买、非法提供信用卡信息罪", "背信运用受托财产罪", "违法运用资金罪","利用未公开信息交易罪", "金融诈骗罪", "集资诈骗罪", "贷款诈骗罪", 
    "票据诈骗罪","金融凭证诈骗罪", "信用证诈骗罪", "信用卡诈骗罪", "有价证券诈骗罪","保险诈骗罪", "危害税收征管罪", "逃税罪", "抗税罪", "逃避追缴欠税罪",
    "骗取出口退税罪", "虚开增值税专用发票、用于骗取出口退税、抵扣税款发票罪","伪造、出售伪造的增值税专用发票罪", "非法出售增值税专用发票罪",
    "非法购买增值税专用发票、购买伪造的增值税专用发票罪","非法制造、出售非法制造的用于骗取出口退税、抵扣税款发票罪","非法制造、出售非法制造的发票罪", 
    "非法出售用于骗取出口退税、抵扣税款发票罪","非法出售发票罪", "虚开发票罪", "持有伪造的发票罪", "侵犯知识产权罪","假冒注册商标罪", "销售假冒注册商标的商品罪",
    "非法制造、销售非法制造的注册商标标识罪", "假冒专利罪", "侵犯著作权罪","销售侵权复制品罪", "侵犯商业秘密罪", "为境外窃取、刺探、收买、非法提供商业秘密罪",
    "扰乱市场秩序罪", "损害商业信誉、商品声誉罪", "虚假广告罪", "串通投标罪","合同诈骗罪", "非法经营罪", "强迫交易罪", "伪造、倒卖伪造的有价票证罪",
    "倒卖车票、船票罪", "非法转让、倒卖土地使用权罪", "提供虚假证明文件罪","出具证明文件重大失实罪", "逃避商检罪", "组织、领导传销活动罪","抢劫罪", "盗窃罪", 
    "诈骗罪", "抢夺罪", "聚众哄抢罪", "侵占罪","职务侵占罪", "挪用资金罪", "挪用特定款物罪", "敲诈勒索罪", "故意毁坏财物罪", "破坏生产经营罪", 
    "拒不支付劳动报酬罪","生产、销售、提供假药罪（原生产、销售假药罪）", "生产、销售、提供劣药罪（原生产、销售劣药罪）",
]

PERSONAL_RIGHTS_CRIMES = [
    "故意杀人罪", "过失致人死亡罪", "故意伤害罪", "过失致人重伤罪", "强奸罪", "非法拘禁罪","绑架罪", "拐卖妇女、儿童罪", "收买被拐卖的妇女、儿童罪", "诬告陷害罪", 
    "强迫劳动罪","非法搜查罪", "非法侵入住宅罪", "侮辱罪", "诽谤罪", "刑讯逼供罪", "暴力取证罪","虐待被监管人罪", "煽动民族仇恨、民族歧视罪", 
    "出版歧视、侮辱少数民族作品罪","非法剥夺公民宗教信仰自由罪", "侵犯少数民族风俗习惯罪", "侵犯通信自由罪","私自开拆、隐匿、毁弃邮件、电报罪", "报复陷害罪", 
    "打击报复会计、统计人员罪","破坏选举罪", "暴力干涉婚姻自由罪", "重婚罪", "破坏军婚罪", "虐待罪", "遗弃罪","拐骗儿童罪", "组织残疾人、儿童乞讨罪", 
    "组织未成年人进行违反治安管理活动罪","组织出卖人体器官罪", "雇用童工从事危重劳动罪", "虐待被监护、看护人罪","负有照护职责人员性侵罪", "抢劫罪", "盗窃罪", 
    "诈骗罪", "抢夺罪", "聚众哄抢罪","侵占罪", "职务侵占罪", "挪用资金罪", "挪用特定款物罪", "敲诈勒索罪","故意毁坏财物罪", "破坏生产经营罪", "拒不支付劳动报酬罪",
    "强制猥亵、侮辱罪（原强制猥亵、侮辱妇女罪）","猥亵儿童罪","聚众阻碍解救被收买的妇女、儿童罪", 
    "侵犯公民个人信息罪（原出售、非法提供公民个人信息罪和非法获取公民个人信息罪）","非法获取公民个人信息罪（已合并）",
]

CASUALTY_CRIMES = [
    "故意伤害罪", "故意杀人罪", "过失致人死亡罪", "过失致人重伤罪", "交通肇事罪","重大飞行事故罪", "重大责任事故罪", "重大劳动安全事故罪", "工程重大安全事故罪",
    "教育设施重大安全事故罪", "大型群众性活动重大安全事故罪"
]

# ====================== 规则定义 ======================
def check_R001(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        ajbh = clean(row["部门受案号"])
        ys = parse_charges(row["移送案由"])
        yst = parse_charges(row["移送其他案由"])
        sx = parse_charges(row["涉嫌案由"])
        sxt = parse_charges(row["涉嫌其他案由"])

        full_case = ys | yst
        person_total = sx | sxt

        if not person_total.issubset(full_case):
            msg = "个人案由超出全案案由范围"
        elif case_person_summary[ajbh]["count"] == 1:
            msg = None if person_total == full_case else "一案一人，全案案由与个人案由不一致"
        else:
            msg = None if case_person_summary[ajbh]["total_set"] == full_case else "一案多人，多人案由之和未覆盖全案案由"

        if msg:
            return build_error_result(
                row, "R001", "审查逮捕案件受理时错填移送罪名", msg, source_table,
                移送案由=clean(row.get("移送案由")),
                移送其他案由=clean(row.get("移送其他案由")),
                涉嫌案由=clean(row.get("涉嫌案由")),
                涉嫌其他案由=clean(row.get("涉嫌其他案由")),
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R001",
            "规则名称": "审查逮捕案件受理时错填移送罪名",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R002(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        bq = clean(row.get("不捕情形"))
        if bq == "其他不捕":
            return build_error_result(
                row, "R002", "是否存在其他不捕情形", "不捕情形为其他不捕", source_table,
                不捕情形=bq,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R002",
            "规则名称": "是否存在其他不捕情形",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R003(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        trial_result = clean(row.get("审结处理结果"))
        if trial_result == "批准逮捕":
            arrest_grounds = clean(row.get("应当逮捕情形"))
            criminal_record = clean(row.get("前科"))
            has_prior_crime = "曾经故意犯罪" in arrest_grounds
            has_criminal_punishment = criminal_record == "刑事处罚"

            if (has_prior_crime and not has_criminal_punishment) or (has_criminal_punishment and not has_prior_crime):
                return build_error_result(
                    row, "R003",
                    '审结结果为批准逮捕案件，应当逮捕情形含"曾经故意犯罪"，前科填"无"',
                    "前科与应当逮捕情形不一致(因前科存在过失犯罪，本条可暂不审核)", source_table,
                    审结处理结果=trial_result,
                    应当逮捕情形=arrest_grounds,
                    前科=criminal_record,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R003",
            "规则名称": '审结结果为批准逮捕案件，应当逮捕情形含"曾经故意犯罪"，前科填"无"',
            "冲突说明": "校验异常：字段缺失"
        }


def check_R004(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        transfer_charge = clean(row.get("移送案由"))
        transfer_other_charge = clean(row.get("移送其他案由"))

        has_official_crime = any(
            crime in transfer_charge or crime in transfer_other_charge
            for crime in OFFICIAL_CRIMES
        )

        investigation_org = clean(row.get("侦查机关名称"))
        transfer_org = clean(row.get("移送单位名称"))

        if not has_official_crime and ("检察院" in investigation_org or "检察院" in transfer_org):
            return build_error_result(
                row, "R004",
                "受理案由为普通刑事犯罪案由，但侦查机关或移送单位错填为检察机关",
                "普通刑事犯罪案由但侦查机关或移送单位错填为检察机关", source_table,
                移送案由=transfer_charge,
                移送其他案由=transfer_other_charge,
                侦查机关名称=investigation_org,
                移送单位名称=transfer_org,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R004",
            "规则名称": "受理案由为普通刑事犯罪案由，但侦查机关或移送单位错填为检察机关",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R005(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        ajbh = clean(row.get("部门受案号"))
        qkaj_jjsjjg = clean(row.get("全案_审结处理结果", ""))
        gr_jjsjjg = clean(row.get("审结处理结果", ""))

        person_count = case_person_summary.get(ajbh, {"count": 1})["count"]

        if person_count == 1:
            if qkaj_jjsjjg != gr_jjsjjg:
                return build_error_result(
                    row, "R005", "逮捕案卡全案审结结果与个人审结结果矛盾",
                    "一案一人，全案_审结处理结果与个人审结处理结果不一致",
                    全案_审结处理结果=qkaj_jjsjjg,
                    审结处理结果=gr_jjsjjg,
                )
        else:
            if (qkaj_jjsjjg == "不批准逮捕" and gr_jjsjjg != "不批准逮捕") or \
               (qkaj_jjsjjg == "不予决定逮捕" and gr_jjsjjg != "不予决定逮捕"):
                return build_error_result(
                    row, "R005", "逮捕案卡全案审结结果与个人审结结果矛盾",
                    "一案多人，全案_审结处理结果与审结处理结果不一致",
                    全案_审结处理结果=qkaj_jjsjjg,
                    审结处理结果=gr_jjsjjg,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R005",
            "规则名称": "逮捕案卡全案审结结果与个人审结结果矛盾",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R006(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        slrq = clean(row.get("受理日期", ""))
        qkaj_sjrq = clean(row.get("全案_审结日期", ""))
        gr_sjrq = clean(row.get("审结日期", ""))

        def parse_date(date_str):
            try:
                return datetime.strptime(date_str, "%Y-%m-%d")
            except:
                try:
                    return datetime.strptime(date_str, "%Y/%m/%d")
                except:
                    return None

        sl_date = parse_date(slrq)
        if not sl_date:
            return None

        for field_name, parsed_date in [("全案_审结日期", parse_date(qkaj_sjrq)), ("审结日期", parse_date(gr_sjrq))]:
            if parsed_date:
                days_diff = (parsed_date - sl_date).days
                if days_diff > 7:
                    return build_error_result(
                        row, "R006", "审查逮捕案件审结日期错填",
                        f"审结日期与受理日期之差大于7天（{days_diff}天）",
                        全案_审结日期=qkaj_sjrq,
                        审结日期=gr_sjrq,
                        受理日期=slrq,
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R006",
            "规则名称": "审查逮捕案件审结日期错填",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R007(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        qkaj_sjrq = clean(row.get("全案_审结日期", ""))
        gr_sjrq = clean(row.get("审结日期", ""))
        szzxhzrq = clean(row.get("收到执行回执日期", ""))
        zxrq = clean(row.get("执行日期", ""))
        zxqk = clean(row.get("执行情况", ""))

        def parse_date(date_str):
            try:
                return datetime.strptime(date_str, "%Y-%m-%d")
            except:
                try:
                    return datetime.strptime(date_str, "%Y/%m/%d")
                except:
                    return None

        current_date = datetime.now().date()
        sjrq_list = []

        qkaj_sj_date = parse_date(qkaj_sjrq)
        if qkaj_sj_date:
            sjrq_list.append(qkaj_sj_date.date())
        gr_sj_date = parse_date(gr_sjrq)
        if gr_sj_date:
            sjrq_list.append(gr_sj_date.date())

        if not sjrq_list:
            return None

        earliest_sjrq = min(sjrq_list)
        days_diff = (current_date - earliest_sjrq).days

        if days_diff > 4:
            if not szzxhzrq or not zxrq or not zxqk:
                return build_error_result(
                    row, "R007", "审查逮捕审结后长期未填录执行情况",
                    f"审结后超过3天未填录执行情况（已逾期{days_diff-4}天）",
                    全案_审结日期=qkaj_sjrq,
                    审结日期=gr_sjrq,
                    收到执行回执日期=szzxhzrq,
                    执行日期=zxrq,
                    执行情况=zxqk,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R007",
            "规则名称": "审查逮捕审结后长期未填录执行情况",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R008(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        trial_charge = clean(row.get("审结案由", ""))
        trial_other_charge = clean(row.get("审结其他案由", ""))
        victim_info = clean(row.get("被害人情况", ""))

        has_personal_rights_crime = any(
            crime in trial_charge or crime in trial_other_charge
            for crime in PERSONAL_RIGHTS_CRIMES
        )

        if has_personal_rights_crime and (victim_info == "无" or not victim_info):
            return build_error_result(
                row, "R008", "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”不应填录（无）",
                "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”不应填录（无）",
                审结案由=trial_charge,
                审结其他案由=trial_other_charge,
                被害人情况=victim_info,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R008",
            "规则名称": "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”不应填录（无）",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R009(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        trial_charge = clean(row.get("审结案由", ""))
        trial_other_charge = clean(row.get("审结其他案由", ""))
        casualty_info = clean(row.get("案件造成伤亡情况", ""))

        has_casualty_crime = any(
            crime in trial_charge or crime in trial_other_charge
            for crime in CASUALTY_CRIMES
        )

        # 检查是否为交通肇事罪且仅造成财产损失的情况
        is_traffic_accident = "交通肇事罪" in trial_charge or "交通肇事罪" in trial_other_charge
        has_casualty = casualty_info != "无" and casualty_info

        # 规则1：应当有案件造成伤亡情况的罪名漏填案件造成伤亡情况
        if has_casualty_crime and (casualty_info == "无" or not casualty_info):
            # 排除交通肇事罪仅造成财产损失的情况
            if not (is_traffic_accident and "财产损失" in casualty_info):
                return build_error_result(
                    row, "R009", "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况",
                    "核实案件造成伤亡情况是否错填",
                    审结案由=trial_charge,
                    审结其他案由=trial_other_charge,
                    案件造成伤亡情况=casualty_info,
                )
        # 规则2：其他案由案件造成案件造成伤亡情况不为"无"或为空时，视为冲突
        elif not has_casualty_crime and has_casualty:
            return build_error_result(
                row, "R009", "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况",
                "案件造成伤亡情况不为空，核实审结案由合理性",
                审结案由=trial_charge,
                审结其他案由=trial_other_charge,
                案件造成伤亡情况=casualty_info,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R009",
            "规则名称": "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R010(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        trial_charge = clean(row.get("审结案由", ""))
        trial_other_charge = clean(row.get("审结其他案由", ""))
        concealment_detail = clean(row.get("掩饰隐瞒犯罪所得、犯罪所得收益罪具体情形", ""))

        has_concealment_crime = "掩饰、隐瞒犯罪所得、犯罪所得收益罪" in trial_charge or "掩饰、隐瞒犯罪所得、犯罪所得收益罪" in trial_other_charge

        if has_concealment_crime and (concealment_detail == "无" or not concealment_detail):
            return build_error_result(
                row, "R010", "审查逮捕案件审结案由为“掩饰隐瞒犯罪所得”时，漏填掩饰隐瞒犯罪所得具体情形",
                "掩饰隐瞒犯罪所得、犯罪所得收益罪具体情形为空或填无",
                审结案由=trial_charge,
                审结其他案由=trial_other_charge,
                掩饰隐瞒犯罪所得犯罪所得收益罪具体情形=concealment_detail,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R010",
            "规则名称": "审查逮捕案件审结案由为“掩饰隐瞒犯罪所得”时，漏填掩饰隐瞒犯罪所得具体情形",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R011(row, case_person_summary=None, source_table="010101表", other_df=None):
    try:
        trial_charge = clean(row.get("审结案由", ""))
        trial_other_charge = clean(row.get("审结其他案由", ""))

        all_charges = trial_charge + trial_other_charge

        for crime in SPECIAL_CRIMES:
            if crime in all_charges:
                return build_error_result(
                    row, "R011", "核查审查逮捕、审查起诉案件特殊罪名",
                    "案由是否错填危害国家安全犯罪或涉恐怖活动犯罪", source_table,
                    审结案由=trial_charge,
                    审结其他案由=trial_other_charge,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R011",
            "规则名称": "核查审查逮捕、审查起诉案件特殊罪名",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R012(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        trial_charge = clean(row.get("审结案由", ""))
        trial_other_charge = clean(row.get("审结其他案由", ""))
        trial_result = clean(row.get("审结处理结果", ""))

        if trial_other_charge == "" and trial_charge == "危险驾驶罪" and trial_result == "批准逮捕":
            return build_error_result(
                row, "R012", "核对审查逮捕案件，单人单案，审结案由为危险驾驶案件被批捕情况",
                "危险驾驶案件，批准逮捕，请核实",
                审结案由=trial_charge,
                审结其他案由=trial_other_charge,
                审结处理结果=trial_result,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R012",
            "规则名称": "核对审查逮捕案件，单人单案，审结案由为危险驾驶案件被批捕情况",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R014(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        trial_charge = clean(row.get("审结案由", ""))
        trial_other_charge = clean(row.get("审结其他案由", ""))
        telecom_flag = clean(row.get("是否利用电信或网络实施犯罪", ""))

        all_charges = trial_charge + trial_other_charge
        
        temp_str = all_charges.replace("危险驾驶罪", "").replace("交通肇事罪", "").replace("、", "").replace("，", "").strip()
        is_only_traffic = temp_str == ""
        
        if is_only_traffic and telecom_flag == "是":
            return build_error_result(
                row, "R014", "危险驾驶、交通肇事类案件填录为“涉及利用电信网络实施犯罪”",
                "危险驾驶或交通肇事类案件错误标记为“是否利用电信或网络实施犯罪”为“是”",
                审结案由=trial_charge,
                审结其他案由=trial_other_charge,
                是否利用电信或网络实施犯罪=telecom_flag,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R014",
            "规则名称": "危险驾驶、交通肇事类案件填录为“涉及利用电信网络实施犯罪”",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R015(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        family_violence = clean(row.get("犯罪涉及校园家庭暴力", ""))
        if family_violence == "涉及家庭暴力":
            return build_error_result(
                row, "R015", "逮捕、起诉案卡是否涉家庭暴力核对",
                "案件涉及家庭暴力，请核对审结案由",
                审结案由=clean(row.get("审结案由")),
                审结其他案由=clean(row.get("审结其他案由")),
                犯罪涉及校园家庭暴力=family_violence,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R015",
            "规则名称": "逮捕、起诉案卡是否涉家庭暴力核对",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R016(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        campus_violence = clean(row.get("犯罪涉及校园家庭暴力", ""))
        if campus_violence == "涉及校园暴力":
            return build_error_result(
                row, "R016", "逮捕、起诉案卡是否涉校园暴力核对",
                "案件涉及校园暴力，请核对审结案由",
                审结案由=clean(row.get("审结案由")),
                审结其他案由=clean(row.get("审结其他案由")),
                犯罪涉及校园家庭暴力=campus_violence,                
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R016",
            "规则名称": "逮捕、起诉案卡是否涉校园暴力核对",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R017(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        if source_table == "010101表":
            # 010101表检查两个字段
            med_dispute = clean(row.get("是否涉医疗纠纷犯罪案件", ""))
            med_beauty = clean(row.get("犯罪案件是否涉及医疗美容犯罪", ""))
            if med_dispute == "是" or med_beauty == "是":
                return build_error_result(
                    row, "R017", "逮捕、起诉案卡涉医核对",
                    "是否涉医犯罪案件，请核对",
                    审结案由=clean(row.get("审结案由")),
                    审结其他案由=clean(row.get("审结其他案由")),
                    是否涉医疗纠纷犯罪案件=med_dispute,
                    犯罪案件是否涉及医疗美容犯罪=med_beauty,
                )
        elif source_table == "020101表":
            # 020101表检查四个字段
            med_dispute = clean(row.get("是否涉医疗纠纷犯罪案件", ""))
            med_beauty = clean(row.get("犯罪案件是否涉及医疗美容犯罪", ""))
            med_device = clean(row.get("犯罪案件是否涉及药品和医用器材安全", ""))
            med_device2 = clean(row.get("是否涉及药品和医用器材安全犯罪", ""))
            if med_dispute == "是" or med_beauty == "是" or med_device == "是" or med_device2 == "是":
                return build_error_result(
                    row, "R017", "逮捕、起诉案卡涉医核对",
                    "是否涉医犯罪案件，请核对",
                    审结案由=clean(row.get("审结案由")),
                    审结其他案由=clean(row.get("审结其他案由")),
                    是否涉医疗纠纷犯罪案件=med_dispute,
                    犯罪案件是否涉及医疗美容犯罪=med_beauty,
                    犯罪案件是否涉及药品和医用器材安全=med_device,
                    是否涉及药品和医用器材安全犯罪=med_device2,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R017",
            "规则名称": "逮捕、起诉案卡涉医核对",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R018(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        if source_table == "010101表":
            # 检查三个字段
            hongkong_macao_taiwan = clean(row.get("涉港澳台侨犯罪案件", ""))
            foreign = clean(row.get("是否涉外案件", ""))
            political_security = clean(row.get("涉政治安全类案件", ""))
            if "涉" in hongkong_macao_taiwan or foreign == "是" or political_security == "是":
                return build_error_result(
                    row, "R018", "逮捕核对涉外、涉侨、涉台、涉港澳案件及涉及政治安全",
                    "涉港澳台侨、涉外或政治安全案件，请核对",
                    审结案由=clean(row.get("审结案由")),
                    审结其他案由=clean(row.get("审结其他案由")),
                    涉港澳台侨犯罪案件=hongkong_macao_taiwan,
                    是否涉外案件=foreign,
                    涉政治安全类案件=political_security,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R018",
            "规则名称": "逮捕核对涉外、涉侨、涉台、涉港澳案件及涉及政治安全",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R019(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        if source_table == "010101表":
            non_public_econ = clean(row.get("是否影响非公有制经济发展案件", ""))
            if non_public_econ == "是":
                return build_error_result(
                    row, "R019", "逮捕案卡是否影响非公经济核对",
                    "是否影响非公有制经济发展案件为是，请核对审结案由",
                    审结案由=clean(row.get("审结案由")),
                    审结其他案由=clean(row.get("审结其他案由")),
                    是否影响非公有制经济发展案件=non_public_econ,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R019",
            "规则名称": "逮捕案卡是否影响非公经济核对",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R022(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        if source_table == "010101表" and other_df is not None:
            # 检查是否为审捕阶段要求提请（移送）逮捕的犯罪嫌疑人
            is_corrective = clean(row.get("是否为审捕阶段要求提请（移送）逮捕的犯罪嫌疑人", ""))
            if is_corrective == "是":
                # 获取嫌疑人编号
                suspect_id = clean(row.get("嫌疑人编号", ""))
                if not suspect_id:
                    return None
                
                # 在020101表中查找对应嫌疑人编号
                found = False
                corrective_in_prosec = ""
                for _, prosec_row in other_df.iterrows():
                    if clean(prosec_row.get("嫌疑人编号", "")) == suspect_id:
                        found = True
                        corrective_in_prosec = clean(prosec_row.get("是否为纠正漏捕的犯罪嫌疑人", ""))
                        break
                
                # 构建错误结果
                if not found:
                    return build_error_result(
                        row, "R022", "在审查逮捕阶段填报为纠正的犯罪嫌疑人，在审查起诉阶段未填为“追捕对象”",
                        "未找到对应一审公诉案件",
                        是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人=is_corrective,
                        是否为纠正漏捕的犯罪嫌疑人="",
                    )
                elif corrective_in_prosec == "否":
                    return build_error_result(
                        row, "R022", "在审查逮捕阶段填报为纠正的犯罪嫌疑人，在审查起诉阶段未填为“追捕对象”",
                        "在审查逮捕阶段填报为纠正的犯罪嫌疑人，在审查起诉阶段未填为“追捕对象”",
                        是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人=is_corrective,
                        是否为纠正漏捕的犯罪嫌疑人=corrective_in_prosec,
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R022",
            "规则名称": "在审查逮捕阶段填报为纠正的犯罪嫌疑人，在审查起诉阶段未填为“追捕对象”",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R023(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表" and other_df is not None:
            # 检查是否为纠正漏捕的犯罪嫌疑人
            is_corrective = clean(row.get("是否为纠正漏捕的犯罪嫌疑人", ""))
            if is_corrective == "是":
                # 获取嫌疑人编号
                suspect_id = clean(row.get("嫌疑人编号", ""))
                if not suspect_id:
                    return None
                
                # 在010101表中查找对应嫌疑人编号
                found = False
                corrective_in_arrest = ""
                for _, arrest_row in other_df.iterrows():
                    if clean(arrest_row.get("嫌疑人编号", "")) == suspect_id:
                        found = True
                        corrective_in_arrest = clean(arrest_row.get("是否为审捕阶段要求提请（移送）逮捕的犯罪嫌疑人", ""))
                        break
                
                # 构建错误结果
                if not found:
                    return build_error_result(
                        row, "R023", "审查起诉阶段填报为“追捕对象”，审查逮捕阶段未填为纠正漏捕的嫌疑人",
                        "未找到对应审查逮捕案件",
                        是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人="",
                        是否为纠正漏捕的犯罪嫌疑人=is_corrective,
                        source_table=source_table
                    )
                elif corrective_in_arrest == "否":
                    return build_error_result(
                        row, "R023", "审查起诉阶段填报为“追捕对象”，审查逮捕阶段未填为纠正漏捕的嫌疑人",
                        "审查起诉阶段填报为“追捕对象”，审查逮捕阶段未填为纠正漏捕的嫌疑人",
                        是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人=corrective_in_arrest,
                        是否为纠正漏捕的犯罪嫌疑人=is_corrective,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R023",
            "规则名称": "审查起诉阶段填报为“追捕对象”，审查逮捕阶段未填为纠正漏捕的嫌疑人",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R024(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        if source_table == "010101表" and other_df is not None:
            # 检查是否为审捕阶段要求提请（移送）逮捕的犯罪嫌疑人
            is_corrective = clean(row.get("是否为审捕阶段要求提请（移送）逮捕的犯罪嫌疑人", ""))
            if is_corrective == "是":
                # 获取嫌疑人编号
                suspect_id = clean(row.get("嫌疑人编号", ""))
                if not suspect_id:
                    return None
                
                # 在020101表中查找对应嫌疑人编号
                found = False
                trial_status = ""
                first_instance_sentence = ""
                first_instance_sentence_term = ""
                
                for _, prosec_row in other_df.iterrows():
                    if clean(prosec_row.get("嫌疑人编号", "")) == suspect_id:
                        found = True
                        trial_status = clean(prosec_row.get("审结处理情况", ""))
                        first_instance_sentence = clean(prosec_row.get("一审宣告刑", ""))
                        first_instance_sentence_term = clean(prosec_row.get("一审宣告刑刑期", ""))
                        break
                
                # 构建错误结果
                if not found:
                    return None
                    
                else:
                    # 检查审结处理情况
                    if any(status in trial_status for status in ["不起诉"]) or trial_status in ["退查后未重报", "同意移送单位撤回", "改变管辖"]:
                        return build_error_result(
                            row, "R024", "核查纠正漏捕后续处理情况是否正确",
                            "纠正漏捕审结处理情况异常，请核对",
                            是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人=is_corrective,
                            审结处理情况=trial_status,
                            一审宣告刑=first_instance_sentence,
                            一审宣告刑刑期=first_instance_sentence_term,
                        )
                    
                    # 检查一审宣告刑
                    if first_instance_sentence in ["免予刑事处罚", "不负刑事责任"] or any(term in first_instance_sentence for term in ["无期", "无罪"]):
                        return build_error_result(
                            row, "R024", "核查纠正漏捕后续处理情况是否正确",
                            "纠正漏捕一审宣告刑异常，请核对",
                            是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人=is_corrective,
                            审结处理情况=trial_status,
                            一审宣告刑=first_instance_sentence,
                            一审宣告刑刑期=first_instance_sentence_term,
                        )
                    
                    # 检查一审宣告刑刑期
                    try:
                        if first_instance_sentence_term and int(first_instance_sentence_term) > 120:
                            return build_error_result(
                                row, "R024", "核查纠正漏捕后续处理情况是否正确",
                                "纠正漏捕一审宣告刑刑期为十年以上，请核对",
                                是否为审捕阶段要求提请_移送_逮捕的犯罪嫌疑人=is_corrective,
                                审结处理情况=trial_status,
                                一审宣告刑=first_instance_sentence,
                                一审宣告刑刑期=first_instance_sentence_term,
                            )
                    except:
                        pass
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R024",
            "规则名称": "核查纠正漏捕后续处理情况是否正确",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R025(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        # 核验“是否黑社会性质组织保护伞”、“犯罪案件是否涉及黑社会性质组织”
        is_protector = clean(row.get("是否黑社会性质组织保护伞", ""))
        is_involve = clean(row.get("犯罪案件是否涉及黑社会性质组织", ""))
        
        if is_protector == "是" or is_protector == "Y" or is_involve == "是":
            return build_error_result(
                row, "R025", "黑势力犯罪情况",
                "黑势力犯罪情况填为是，请核实",
                是否黑社会性质组织保护伞=is_protector,
                犯罪案件是否涉及黑社会性质组织=is_involve,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R025",
            "规则名称": "黑势力犯罪情况",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R026(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        # 核验“是否恶势力保护伞”、“犯罪案件是否涉及恶势力”
        is_protector = clean(row.get("是否恶势力保护伞", ""))
        is_involve = clean(row.get("犯罪案件是否涉及恶势力", ""))
        
        if is_protector == "是" or is_involve == "是":
            return build_error_result(
                row, "R026", "恶势力犯罪情况",
                "恶势力犯罪情况填为是，请核实",
                是否恶势力保护伞=is_protector,
                犯罪案件是否涉及恶势力=is_involve,
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R026",
            "规则名称": "恶势力犯罪情况",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R027(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        special_work = clean(row.get("专项活动重点工作", ""))
        target_values = ["涉及农资", "涉及食品安全", "涉及药品和医用器材安全", "涉及食用农产品安全"]
        
        if special_work in target_values:
            return build_error_result(
                row, "R027", "审查逮捕、审查起诉错填涉农资犯罪、涉食药品犯罪",
                "核实是否错填涉农资犯罪、涉食药品犯罪",
                专项活动重点工作=special_work,
                source_table=source_table
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R027",
            "规则名称": "审查逮捕、审查起诉错填涉农资犯罪、涉食药品犯罪",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R028(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_status = clean(row.get("审结处理情况", ""))
            if trial_status == "附条件不起诉":
                trial_charge = clean(row.get("审结案由", ""))
                trial_other_charge = clean(row.get("审结其他案由", ""))
                age = clean(row.get("作案时年龄", ""))
                
                # 定义四、五、六章犯罪列表
                CHAPTER_4_5_6_CRIMES = [
                    "故意杀人罪", "过失致人死亡罪", "故意伤害罪", "过失致人重伤罪", "强奸罪",
                    "强制猥亵、侮辱罪（原强制猥亵、侮辱妇女罪）", "猥亵儿童罪", "非法拘禁罪", "绑架罪",
                    "拐卖妇女、儿童罪", "收买被拐卖的妇女、儿童罪", "聚众阻碍解救被收买的妇女、儿童罪",
                    "诬告陷害罪", "强迫劳动罪", "非法搜查罪", "非法侵入住宅罪", "侮辱罪",
                    "诽谤罪", "刑讯逼供罪", "暴力取证罪", "虐待被监管人罪", "煽动民族仇恨、民族歧视罪",
                    "出版歧视、侮辱少数民族作品罪", "非法剥夺公民宗教信仰自由罪", "侵犯少数民族风俗习惯罪",
                    "侵犯通信自由罪", "私自开拆、隐匿、毁弃邮件、电报罪", "报复陷害罪",
                    "打击报复会计、统计人员罪", "破坏选举罪", "暴力干涉婚姻自由罪", "重婚罪",
                    "破坏军婚罪", "虐待罪", "遗弃罪", "拐骗儿童罪", "组织残疾人、儿童乞讨罪",
                    "侵犯公民个人信息罪（原出售、非法提供公民个人信息罪和非法获取公民个人信息罪）",
                    "非法获取公民个人信息罪（已合并）", "组织未成年人进行违反治安管理活动罪",
                    "组织出卖人体器官罪", "雇用童工从事危重劳动罪", "虐待被监护、看护人罪",
                    "负有照护职责人员性侵罪", "抢劫罪", "盗窃罪", "诈骗罪", "抢夺罪",
                    "聚众哄抢罪", "侵占罪", "职务侵占罪", "挪用资金罪", "挪用特定款物罪",
                    "敲诈勒索罪", "故意毁坏财物罪", "破坏生产经营罪", "拒不支付劳动报酬罪",
                    "妨害公务罪", "煽动暴力抗拒法律实施罪", "招摇撞骗罪",
                    "伪造、变造、买卖国家机关公文、证件、印章罪",
                    "盗窃、抢夺、毁灭国家机关公文、证件、印章罪",
                    "伪造公司、企业、事业单位、人民团体印章罪",
                    "伪造、变造、买卖身份证件罪（原伪造、变造居民身份证罪）",
                    "非法生产、买卖警用装备罪", "非法获取国家秘密罪",
                    "非法持有国家绝密、机密文件、资料、物品罪",
                    "非法生产、销售专用间谍器材、窃听、窃照专用器材罪（原非法生产、销售间谍专用器材罪）",
                    "非法使用窃听、窃照专用器材罪", "非法侵入计算机信息系统罪",
                    "破坏计算机信息系统罪", "扰乱无线电通讯管理秩序罪", "聚众扰乱社会秩序罪",
                    "聚众冲击国家机关罪", "聚众扰乱公共场所秩序、交通秩序罪", "投放虚假危险物质罪",
                    "编造、故意传播虚假恐怖信息罪", "聚众斗殴罪", "寻衅滋事罪",
                    "组织、领导、参加黑社会性质组织罪", "入境发展黑社会组织罪",
                    "包庇、纵容黑社会性质组织罪", "传授犯罪方法罪", "非法集会、游行、示威罪",
                    "非法携带武器、管制刀具、爆炸物参加集会、游行、示威罪", "破坏集会、游行、示威罪",
                    "侮辱国旗、国徽、国歌罪（原侮辱国旗、国徽罪）",
                    "组织、利用会道门、邪教组织、利用迷信破坏法律实施罪",
                    "组织、利用会道门、邪教组织、利用迷信致人重伤、死亡罪（原组织、利用会道门、邪教组织、利用迷信致人死亡罪）",
                    "聚众淫乱罪", "引诱未成年人聚众淫乱罪",
                    "盗窃、侮辱、故意毁坏尸体、尸骨、骨灰罪（原盗窃、侮辱尸体罪）",
                    "赌博罪", "故意延误投递邮件罪", "开设赌场罪",
                    "非法获取计算机信息系统数据、非法控制计算机信息系统罪",
                    "提供侵入、非法控制计算机信息系统程序、工具罪", "使用虚假身份证件、盗用身份证件罪",
                    "组织考试作弊罪", "非法出售、提供试题、答案罪", "代替考试罪",
                    "拒不履行信息网络安全管理义务罪", "非法利用信息网络罪", "帮助信息网络犯罪活动罪",
                    "扰乱国家机关工作秩序罪", "组织、资助非法聚集罪", "编造、故意传播虚假信息罪",
                    "袭警罪", "冒名顶替罪", "高空抛物罪", "催收非法债务罪", "侵害英雄烈士名誉、荣誉罪",
                    "组织参与国（境）外赌博罪", "妨害司法罪", "伪证罪",
                    "辩护人、诉讼代理人毁灭证据、伪造证据、妨害作证罪", "妨害作证罪",
                    "帮助毁灭、伪造证据罪", "打击报复证人罪", "扰乱法庭秩序罪", "窝藏、包庇罪",
                    "拒绝提供间谍犯罪、恐怖主义犯罪、极端主义犯罪证据罪（原拒绝提供间谍犯罪证据罪）",
                    "掩饰、隐瞒犯罪所得、犯罪所得收益罪", "拒不执行判决、裁定罪",
                    "非法处置查封、扣押、冻结的财产罪", "破坏监管秩序罪", "脱逃罪",
                    "劫夺被押解人员罪", "组织越狱罪", "暴动越狱罪", "聚众持械劫狱罪", "虚假诉讼罪",
                    "泄露不应公开的案件信息罪", "披露、报道不应公开的案件信息罪", "妨害国边境管理罪",
                    "组织他人偷越国(边)境罪", "骗取出境证件罪", "提供伪造、变造的出入境证件罪",
                    "出售出入境证件罪", "运送他人偷越国(边)境罪", "偷越国(边)境罪",
                    "破坏界碑、界桩罪", "破坏永久性测量标志罪", "妨害文物管理罪", "故意损毁文物罪",
                    "故意损毁名胜古迹罪", "过失损毁文物罪", "非法向外国人出售、赠送珍贵文物罪",
                    "倒卖文物罪", "非法出售、私赠文物藏品罪", "盗掘古文化遗址、古墓葬罪",
                    "盗掘古人类化石、古脊椎动物化石罪", "抢夺、窃取国有档案罪",
                    "擅自出卖、转让国有档案罪", "危害公共卫生罪", "妨害传染病防治罪",
                    "传染病菌种、毒种扩散罪", "妨害国境卫生检疫罪", "非法组织卖血罪", "强迫卖血罪",
                    "非法采集、供应血液、制作、供应血液制品罪", "采集、供应血液、制作、供应血液制品事故罪",
                    "医疗事故罪", "非法行医罪", "非法进行节育手术罪", "妨害动植物防疫、检疫罪",
                    "非法采集人类遗传资源、走私人类遗传资源材料罪", "非法植入基因编辑、克隆胚胎罪",
                    "破坏环境资源保护罪", "污染环境罪", "非法处置进口的固体废物罪", "擅自进口固体废物罪",
                    "非法捕捞水产品罪",
                    "危害珍贵、濒危野生动物罪（原非法猎捕、杀害珍贵、濒危野生动物罪和非法收购、运输、出售珍贵、濒危野生动物、珍贵、濒危野生动物制品罪）",
                    "非法收购、运输、出售珍贵、濒危野生动物、珍贵、濒危野生动物制品罪（已合并）",
                    "非法狩猎罪", "非法占用农用地罪", "非法采矿罪", "破坏性采矿罪",
                    "危害国家重点保护植物罪（原非法采伐、毁坏国家重点保护植物罪和非法收购、运输、加工、出售国家重点保护植物、国家重点保护植物制品罪）",
                    "盗伐林木罪", "滥伐林木罪", "非法收购、运输盗伐、滥伐的林木罪",
                    "非法收购、运输、加工、出售国家重点保护植物、国家重点保护植物制品罪（已合并）",
                    "非法猎捕、收购、运输、出售陆生野生动物罪", "破坏自然保护地罪",
                    "非法引进、释放、丢弃外来入侵物种罪", "走私、贩卖、运输、制造毒品罪",
                    "走私、贩卖、运输、制造毒品罪", "非法持有毒品罪", "包庇毒品犯罪分子罪",
                    "窝藏、转移、隐瞒毒品、毒赃罪",
                    "非法生产、买卖、运输制毒物品、走私制毒物品罪（原走私制毒物品罪和非法买卖制毒物品罪）",
                    "非法买卖制毒物品罪（已合并）", "非法种植毒品原植物罪",
                    "非法买卖、运输、携带、持有毒品原植物种子、幼苗罪", "引诱、教唆、欺骗他人吸毒罪",
                    "强迫他人吸毒罪", "容留他人吸毒罪", "非法提供麻醉药品、精神药品罪", "妨害兴奋剂管理罪",
                    "组织、强迫、引诱、容留、介绍卖淫罪", "组织卖淫罪", "强迫卖淫罪", "协助组织卖淫罪",
                    "引诱、容留、介绍卖淫罪", "引诱幼女卖淫罪", "传播性病罪", "嫖宿幼女罪（已删除）",
                    "制作、贩卖、传播淫秽物品罪", "制作、复制、出版、贩卖、传播淫秽物品牟利罪",
                    "为他人提供书号出版淫秽书刊罪", "传播淫秽物品罪", "组织播放淫秽音像制品罪", "组织淫秽表演罪"
                ]
                
                # 检查是否在四、五、六章犯罪中
                in_chapter_crimes = any(
                    crime in trial_charge or crime in trial_other_charge
                    for crime in CHAPTER_4_5_6_CRIMES
                )
                
                # 检查年龄是否小于18
                is_minor = False
                try:
                    if age and float(age) < 18:
                        is_minor = True
                except:
                    pass
                
                # 冲突条件：不在四、五、六章犯罪中，或者年龄大于等于18
                if not in_chapter_crimes or not is_minor:
                    return build_error_result(
                        row, "R028", "非未成年人附条件不起诉，或附条件不起诉罪名是否为四、五、六章犯罪",
                        "附条件不起诉，核对年龄和审结案由",
                        审结处理情况=trial_status,
                        审结案由=trial_charge,
                        审结其他案由=trial_other_charge,
                        作案时年龄=age,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R028",
            "规则名称": "非未成年人附条件不起诉，或附条件不起诉罪名是否为四、五、六章犯罪",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R029(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        ajbh = clean(row.get("部门受案号", ""))
        ys = parse_charges(row.get("全案_移送案由", ""))
        yst = parse_charges(row.get("全案_移送其他案由", ""))
        sx = parse_charges(row.get("移诉案由", ""))
        sxt = parse_charges(row.get("移诉其他案由", ""))

        full_case = ys | yst
        person_total = sx | sxt

        if not person_total.issubset(full_case):
            msg = "个人案由超出全案案由范围"
        elif case_person_summary.get(ajbh, {"count": 1})["count"] == 1:
            msg = None if person_total == full_case else "一案一人，全案案由与个人案由不一致"
        else:
            msg = None if case_person_summary.get(ajbh, {"total_set": set()})["total_set"] == full_case else "一案多人，多人案由之和未覆盖全案案由"

        if msg:
            return build_error_result(
                row, "R029", "一审公诉案件受理时错填移送罪名", msg, source_table,
                全案_移送案由=clean(row.get("全案_移送案由")),
                全案_移送其他案由=clean(row.get("全案_移送其他案由")),
                移诉案由=clean(row.get("移诉案由")),
                移诉其他案由=clean(row.get("移诉其他案由")),
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R029",
            "规则名称": "一审公诉案件受理时错填移送罪名",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R030(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            ys意见 = clean(row.get("全案_移送意见", ""))
            if ys意见 == "移送不起诉":
                return build_error_result(
                    row, "R030", "受理移送不起诉是否错填",
                    "全案_移送意见是否漏填或错填",
                    全案_移送意见=ys意见,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R030",
            "规则名称": "受理移送不起诉是否错填",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R031(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 检查审结案由和审结其他案由
            trial_charge = clean(row.get("审结案由", ""))
            trial_other_charge = clean(row.get("审结其他案由", ""))
            
            # 检查是否包含经济类、侵财类犯罪
            has_economic_crime = any(
                crime in trial_charge or crime in trial_other_charge
                for crime in ECONOMIC_CRIMES
            )
            
            # 检查追赃挽损情况
            zang损失 = clean(row.get("案件是否涉及追赃挽损", ""))
            if has_economic_crime and (zang损失 == "否" or zang损失 == ""):
                return build_error_result(
                    row, "R031", "经济类、侵财类犯罪核对'涉及追赃挽损情况'",
                    "经济类、侵财类犯罪，核对追赃挽损情况是否漏填",
                    审结案由=trial_charge,
                    审结其他案由=trial_other_charge,
                    案件是否涉及追赃挽损=zang损失,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R031",
            "规则名称": "经济类、侵财类犯罪核对'涉及追赃挽损情况'",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R032(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 检查审结案由和审结其他案由
            trial_charge = clean(row.get("审结案由", ""))
            trial_other_charge = clean(row.get("审结其他案由", ""))
            
            # 检查是否包含侵犯公民人身权利、民主权利及财产类犯罪
            has_personal_rights_crime = any(
                crime in trial_charge or crime in trial_other_charge
                for crime in PERSONAL_RIGHTS_CRIMES
            )
            
            # 检查被害人情况
            victim_info = clean(row.get("被害人情况", ""))
            if has_personal_rights_crime and (victim_info == "无" or victim_info == ""):
                return build_error_result(
                    row, "R032", "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”不应填录（无）",
                    "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”漏填或错填",
                    审结案由=trial_charge,
                    审结其他案由=trial_other_charge,
                    被害人情况=victim_info,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R032",
            "规则名称": "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”不应填录（无）",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R033(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        trial_charge = clean(row.get("全案_审结案由", ""))
        trial_other_charge = clean(row.get("全案_审结其他案由", ""))
        casualty_info = clean(row.get("案件造成伤亡情况", ""))

        has_casualty_crime = any(
            crime in trial_charge or crime in trial_other_charge
            for crime in CASUALTY_CRIMES
        )

        # 检查是否为交通肇事罪且仅造成财产损失的情况
        is_traffic_accident = "交通肇事罪" in trial_charge or "交通肇事罪" in trial_other_charge
        has_casualty = casualty_info != "无" and casualty_info

        # 规则1：应当有案件造成伤亡情况的罪名漏填案件造成伤亡情况
        if has_casualty_crime and (casualty_info == "无" or not casualty_info):
            # 排除交通肇事罪仅造成财产损失的情况
            if not (is_traffic_accident and "财产损失" in casualty_info):
                return build_error_result(
                    row, "R033", "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况",
                    "核实案件造成伤亡情况是否错填",
                    全案_审结案由=trial_charge,
                    全案_审结其他案由=trial_other_charge,
                    案件造成伤亡情况=casualty_info,
                    source_table=source_table
                )
        # 规则2：其他案由案件造成案件造成伤亡情况不为"无"或为空时，视为冲突
        elif not has_casualty_crime and has_casualty:
            return build_error_result(
                row, "R033", "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况",
                "案件造成伤亡情况不为空，核实审结案由合理性",
                全案_审结案由=trial_charge,
                全案_审结其他案由=trial_other_charge,
                案件造成伤亡情况=casualty_info,
                source_table=source_table
            )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R033",
            "规则名称": "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R034(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 获取受理日期
            slrq = clean(row.get("全案_受理日期", ""))
            # 获取审结日期
            qkaj_sjrq = clean(row.get("全案_审结日期", ""))
            gr_sjrq = clean(row.get("审结日期", ""))

            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None

            # 解析受理日期
            sl_date = parse_date(slrq)
            if not sl_date:
                return None

            # 计算距离昨天的天数
            yesterday = datetime.now().date()
            days_diff = (yesterday - sl_date.date()).days

            # 检查是否超过366天且未审结
            if days_diff > 366 and (not qkaj_sjrq or not gr_sjrq):
                return build_error_result(
                    row, "R034", "一审公诉案件受理后超过一年未审结",
                    "一审公诉案件受理后超过一年未审结",
                    全案_受理日期=slrq,
                    全案_审结日期=qkaj_sjrq,
                    审结日期=gr_sjrq,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R034",
            "规则名称": "一审公诉案件受理后超过一年未审结",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R035(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 检查审结情况
           
            sjcl情况 = clean(row.get("审结处理情况", ""))
            
            # 只有当审结情况或审结处理情况等于起诉时才检查
            if  sjcl情况 != "起诉":
                return None
            
            # 获取审结日期
            qkaj_sjrq = clean(row.get("全案_审结日期", ""))
            gr_sjrq = clean(row.get("审结日期", ""))
            # 获取一审判决日期
            qkaj_pjrq = clean(row.get("全案_一审判决日期", ""))
            gr_pjrq = clean(row.get("一审判决日期", ""))

            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None

            # 解析审结日期
            sj_date = None
            if qkaj_sjrq:
                sj_date = parse_date(qkaj_sjrq)
            elif gr_sjrq:
                sj_date = parse_date(gr_sjrq)
            
            if not sj_date:
                return None

            # 计算距离昨天的天数
            yesterday = datetime.now().date()
            days_diff = (yesterday - sj_date.date()).days

            # 检查是否超过366天且无判决
            if days_diff > 366 and (not qkaj_pjrq or not gr_pjrq):
                return build_error_result(
                    row, "R035", "一审公诉案件审结后超过一年无判决",
                    "一审公诉案件审结后超过一年无判决",
                    审结处理情况=sjcl情况,
                    全案_审结日期=qkaj_sjrq,
                    审结日期=gr_sjrq,
                    全案_一审判决日期=qkaj_pjrq,
                    一审判决日期=gr_pjrq,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R035",
            "规则名称": "一审公诉案件审结后超过一年无判决",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R036(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 获取庭审程序
            tsxc = clean(row.get("庭审程序", ""))
            
            # 只有庭审程序为速裁程序时才检查
            if tsxc != "速裁程序":
                return None
            
            # 获取一审宣告刑刑期
            xgxq = clean(row.get("一审宣告刑刑期", ""))
            # 获取日期
            slrq = clean(row.get("全案_受理日期", ""))
            sjrq = clean(row.get("全案_审结日期", ""))

            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None

            def parse_xgxq(xgxq_str):
                try:
                    # 提取数字
                    import re
                    match = re.search(r'\d+', xgxq_str)
                    if match:
                        return int(match.group())
                    return None
                except:
                    return None

            # 解析日期
            sl_date = parse_date(slrq)
            sj_date = parse_date(sjrq)
            
            if not sl_date or not sj_date:
                return None

            # 计算天数差
            days_diff = (sj_date.date() - sl_date.date()).days

            # 解析一审宣告刑刑期
            xgxq_value = parse_xgxq(xgxq)

            # 检查两种情况
            if (xgxq_value is None or xgxq_value <= 12) and days_diff > 10:
                # 情况1：刑期小于等于12个月或为空，超过10天
                return build_error_result(
                    row, "R036", "速裁程序办案时限十天，可能判处有期徒刑超过一年的，可以延长十五日",
                    "核实速裁程序办案期限是否超期",
                    庭审程序=tsxc,
                    一审宣告刑刑期=xgxq,
                    全案_受理日期=slrq,
                    全案_审结日期=sjrq,
                    办案天数=days_diff,
                    source_table=source_table
                )
            elif (xgxq_value is None or xgxq_value > 12) and days_diff > 15:
                # 情况2：刑期大于12个月或为空，超过15天
                return build_error_result(
                    row, "R036", "速裁程序办案时限十天，可能判处有期徒刑超过一年的，可以延长十五日",
                    "核实速裁程序办案期限是否超期",
                    庭审程序=tsxc,
                    一审宣告刑刑期=xgxq,
                    全案_受理日期=slrq,
                    全案_审结日期=sjrq,
                    办案天数=days_diff,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R036",
            "规则名称": "速裁程序办案时限十天，可能判处有期徒刑超过一年的，可以延长十五日",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R037(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 获取相关字段
            sflajd = clean(row.get("是否立案监督对象", ""))
            sfzss = clean(row.get("是否为追诉对象", ""))
            sfjzlfb = clean(row.get("是否为纠正漏捕的犯罪嫌疑人", ""))

            # 统计填写为"是"的数量
            yes_count = 0
            if sflajd == "是":
                yes_count += 1
            if sfzss == "是":
                yes_count += 1
            if sfjzlfb == "是":
                yes_count += 1

            # 任意两个同时填写为"是"则视为冲突
            if yes_count >= 2:
                return build_error_result(
                    row, "R037", "同一犯罪嫌疑人既立案监督又纠正漏捕又纠正漏诉",
                    "同一犯罪嫌疑人，立案监督、纠正漏捕、纠正漏诉填写超过一个",
                    是否立案监督对象=sflajd,
                    是否为追诉对象=sfzss,
                    是否为纠正漏捕的犯罪嫌疑人=sfjzlfb,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R037",
            "规则名称": "同一犯罪嫌疑人既立案监督又纠正漏捕又纠正漏诉",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R038(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            # 获取一次退查相关字段
            yctcdr = clean(row.get("全案_一次退查日期", ""))
            yczbrq = clean(row.get("全案_一次重报日期", ""))
            yctchwczs = clean(row.get("一次退查后未重报事由", ""))
            
            # 获取二次退查相关字段
            ectcdr = clean(row.get("全案_二次退查日期", ""))
            eczbrq = clean(row.get("全案_二次重报日期", ""))
            ectchwczs = clean(row.get("二次退查后未重报事由", ""))

            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None

            # 当前系统日期
            today = datetime.now().date()

            # 检查第一种情况：一次退查
            if yctcdr:
                yctc_date = parse_date(yctcdr)
                if yctc_date:
                    days_diff = (today - yctc_date.date()).days
                    if days_diff > 31 and not yczbrq and not yctchwczs:
                        return build_error_result(
                            row, "R038", "一审公诉案件退查后长期未重报且未填报“退查未重报”",
                            "一审公诉案件退查后未重报且未填写“退查未重报事由”",
                            全案_一次退查日期=yctcdr,
                            全案_一次重报日期=yczbrq,
                            一次退查后未重报事由=yctchwczs,
                            全案_二次退查日期=ectcdr,
                            全案_二次重报日期=eczbrq,
                            二次退查后未重报事由=ectchwczs,
                            退查天数=days_diff,
                            source_table=source_table
                        )

            # 检查第二种情况：二次退查
            if ectcdr:
                ectc_date = parse_date(ectcdr)
                if ectc_date:
                    days_diff = (today - ectc_date.date()).days
                    if days_diff > 31 and not eczbrq and not ectchwczs:
                        return build_error_result(
                            row, "R038", "一审公诉案件退查后长期未重报且未填报“退查未重报”",
                            "一审公诉案件退查后未重报且未填写“退查未重报事由”",
                            全案_二次退查日期=ectcdr,
                            全案_二次重报日期=eczbrq,
                            二次退查后未重报事由=ectchwczs,
                            全案_一次退查日期=yctcdr,
                            全案_一次重报日期=yczbrq,
                            一次退查后未重报事由=yctchwczs,
                            退查天数=days_diff,
                            source_table=source_table
                        )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R038",
            "规则名称": "一审公诉案件退查后长期未重报且未填报'退查未重报'",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R039(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_status = clean(row.get("审结处理情况", ""))
            全案认罪认罚 = clean(row.get("全案_审查起诉阶段检察机关适用认罪认罚", ""))
            个人认罪认罚 = clean(row.get("审查起诉阶段检察机关适用认罪认罚", ""))
            
            if (trial_status == "绝对不起诉" or trial_status == "其他证据不足不起诉") and (全案认罪认罚 == "是" or 个人认罪认罚 == "是"):
                return build_error_result(
                    row, "R039", "核对是否存在存疑不诉、绝对不诉案件适用认罪认罚",
                    "存疑不诉、绝对不诉案件适用认罪认罚，请核对",
                    审结处理情况=trial_status,
                    全案_审查起诉阶段检察机关适用认罪认罚=全案认罪认罚,
                    审查起诉阶段检察机关适用认罪认罚=个人认罪认罚,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R039",
            "规则名称": "核对是否存在存疑不诉、绝对不诉案件适用认罪认罚",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R040(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_status = clean(row.get("审结处理情况", ""))
            全案认罪认罚 = clean(row.get("全案_审查起诉阶段检察机关适用认罪认罚", ""))
            个人认罪认罚 = clean(row.get("审查起诉阶段检察机关适用认罪认罚", ""))
            
            if (trial_status == "其他情节轻微不起诉" or trial_status == "附条件不起诉" or trial_status == "刑事和解不起诉") and (全案认罪认罚 == "否" or 个人认罪认罚 == "否"):
                return build_error_result(
                    row, "R040", "核对相对不诉但未适用认罪认罚",
                    "存疑不诉、绝对不诉案件适用认罪认罚，请核对",
                    审结处理情况=trial_status,
                    全案_审查起诉阶段检察机关适用认罪认罚=全案认罪认罚,
                    审查起诉阶段检察机关适用认罪认罚=个人认罪认罚,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R040",
            "规则名称": "核对相对不诉但未适用认罪认罚",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R041(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            xyrbh = clean(row.get("嫌疑人编号", ""))
            change_jurisdiction = clean(row.get("是否他院受理审查起诉后改变管辖", ""))
            
            xyrbh_count = case_person_summary.get("xyrbh_count", {})
            if xyrbh and xyrbh_count.get(xyrbh, 1) > 1 and (change_jurisdiction == "否" or change_jurisdiction == ""):
                return build_error_result(
                    row, "R041", "是否他院受理后改变管辖漏填",
                    "自侦案件是否他院受理审查起诉后改变管辖可能漏填，请核对",
                    嫌疑人编号=xyrbh,
                    是否他院受理审查起诉后改变管辖=change_jurisdiction,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R041",
            "规则名称": "是否他院受理后改变管辖漏填",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R042(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            amount_fields = [
                "保证金（万元）",
                "审结数额（万元）",
                "认定造成经济损失（万元）",
                "挽回经济损失（万元）",
                "侦（调）查机关追缴款物（万元）",
                "检察机关自行查扣冻（万元）",
                "检察机关发现并由其他机关查扣冻（万元）",
                "审定数额（万元）",
                "洗钱数额（万元）",
                "掩饰隐瞒犯罪所得、犯罪所得收益数额（万元）",
                "捕诉环节犯罪嫌疑人主动退赔（万元）",
                "一审罚金数额（万元）",
                "二审罚金数额（万元）"
            ]
            
            conflict_fields = {}
            count_gt_499 = 0
            
            for field in amount_fields:
                value = row.get(field)
                try:
                    num_value = float(value) if pd.notna(value) else 0
                    if num_value > 499:
                        count_gt_499 += 1
                        conflict_fields[field] = num_value
                except:
                    pass
            
            if count_gt_499 == 1:
                return build_error_result(
                    row, "R042", "核对一审公诉案件审结、追赃挽损填录异常问题",
                    "核对金额是否填错，单位为万元",
                    source_table=source_table,
                    **conflict_fields
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R042",
            "规则名称": "核对一审公诉案件审结、追赃挽损填录异常问题",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R043(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            withdraw_date = clean(row.get("撤回起诉日期", ""))
            process_date = clean(row.get("撤回起诉后处理日期", ""))
            
            if withdraw_date:
                try:
                    withdraw_date_obj = pd.to_datetime(withdraw_date)
                    current_date = pd.to_datetime("today")
                    days_diff = (current_date - withdraw_date_obj).days
                    
                    if days_diff > 31 and not process_date:
                        return build_error_result(
                            row, "R043", "撤回起诉后一个月案件无处理结果",
                            "撤回起诉后一个月案件无处理结果",
                            撤回起诉日期=withdraw_date,
                            撤回起诉后处理日期=process_date,
                            source_table=source_table
                        )
                except:
                    pass
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R043",
            "规则名称": "撤回起诉后一个月案件无处理结果",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R044(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            qa_sj_date = clean(row.get("全案_审结日期", ""))
            sj_date = clean(row.get("审结日期", ""))
            
            if not qa_sj_date or not sj_date:
                return None
            
            ys = parse_charges(row.get("移诉案由", ""))
            yst = parse_charges(row.get("移诉其他案由", ""))
            sx = parse_charges(row.get("审结案由", ""))
            sxt = parse_charges(row.get("审结其他案由", ""))
            
            ys_total = ys | yst
            sx_total = sx | sxt
            
            change_type = clean(row.get("改变移送审查起诉认定情形", ""))
            has_change_定性 = "改变定性" in change_type
            
            if ys_total == sx_total:
                if has_change_定性:
                    return build_error_result(
                        row, "R044", "审查起诉改变侦(调)查认定的情形漏填",
                        "个人移送罪名与审结罪名一致，核查错填改变定性",
                        移诉案由=clean(row.get("移诉案由")),
                        移诉其他案由=clean(row.get("移诉其他案由")),
                        审结案由=clean(row.get("审结案由")),
                        审结其他案由=clean(row.get("审结其他案由")),
                        改变移送审查起诉认定情形=change_type,
                        source_table=source_table
                    )
            else:
                if not has_change_定性:
                    return build_error_result(
                        row, "R044", "审查起诉改变侦(调)查认定的情形漏填",
                        "个人移送罪名与审结罪名不一致，核查漏填改变定性",
                        移诉案由=clean(row.get("移诉案由")),
                        移诉其他案由=clean(row.get("移诉其他案由")),
                        审结案由=clean(row.get("审结案由")),
                        审结其他案由=clean(row.get("审结其他案由")),
                        改变移送审查起诉认定情形=change_type,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R044",
            "规则名称": "审查起诉改变侦(调)查认定的情形漏填",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R045(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            withdrawal_date = clean(row.get("撤回起诉日期", ""))
            first_instance_sentence = clean(row.get("一审宣告刑", ""))
            
            if withdrawal_date:
                return build_error_result(
                    row, "R045", "核对无罪、撤回起诉、不负刑事责任案件填报情况",
                    "核对撤回起诉案件是否存在跨年填报情况",
                    撤回起诉日期=withdrawal_date,
                    source_table=source_table
                )
            
            if first_instance_sentence == "免予刑事处罚" or first_instance_sentence == "不负刑事责任" or "无罪" in first_instance_sentence:
                return build_error_result(
                    row, "R045", "核对无罪、撤回起诉、不负刑事责任案件填报情况，与50条一并核对。",
                    "核对无罪、不予刑事处罚、不负刑事责任案件填报情况，与50条一并核对。",
                    一审宣告刑=first_instance_sentence,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R045",
            "规则名称": "核对无罪、撤回起诉、不负刑事责任案件填报情况，与50条一并核对。",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R046(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_status = clean(row.get("审结处理情况", ""))
            prosecution_date = clean(row.get("提起公诉日期", ""))
            
            if trial_status == "起诉" and not prosecution_date:
                return build_error_result(
                    row, "R046", "漏填提起公诉日期",
                    "漏填提起公诉日期",
                    审结处理情况=trial_status,
                    提起公诉日期=prosecution_date,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R046",
            "规则名称": "漏填提起公诉日期",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R047(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            receive_date = clean(row.get("一审判决书收到日期", ""))
            
            if not receive_date:
                return None
            
            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None
            
            receive_date_parsed = parse_date(receive_date)
            if not receive_date_parsed:
                return None
            
            current_date = datetime.now()
            days_diff = (current_date - receive_date_parsed).days
            
            if days_diff > 11:
                effective_date = clean(row.get("一审判决生效日期", ""))
                appeal_date = clean(row.get("上诉日期", ""))
                protest_date = clean(row.get("提出抗诉日期", ""))
                
                if not effective_date and not appeal_date and not protest_date:
                    return build_error_result(
                        row, "R047", "一审判决后未上诉抗诉长期不填报生效日期",
                        "收到判决书十日后，漏填判决生效日期、上诉日期或抗诉日期",
                        一审判决书收到日期=receive_date,
                        一审判决生效日期=effective_date,
                        上诉日期=appeal_date,
                        提出抗诉日期=protest_date,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R047",
            "规则名称": "一审判决后未上诉抗诉长期不填报生效日期",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R048(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_status = clean(row.get("审结处理情况", ""))
            
            if trial_status != "起诉":
                return None
            
            case_effective_date = clean(row.get("全案_一审判决生效日期", ""))
            
            if case_effective_date:
                person_judgment_date = clean(row.get("一审判决日期", ""))
                person_charge = clean(row.get("一审判决罪名/案由", ""))
                person_sentence = clean(row.get("一审宣告刑", ""))
                
                if not person_judgment_date or not person_charge or not person_sentence:
                    return build_error_result(
                        row, "R048", "一审公诉填报案件裁判结果，人员裁判结果未填报",
                        "一审公诉案卡填有裁判结果，漏填人卡裁判结果",
                        全案_一审判决生效日期=case_effective_date,
                        一审判决日期=person_judgment_date,
                        一审判决罪名_案由=person_charge,
                        一审宣告刑=person_sentence,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R048",
            "规则名称": "一审公诉填报案件裁判结果，人员裁判结果未填报",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R049(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            protest_date = clean(row.get("提出抗诉日期", ""))
            withdraw_protest_date = clean(row.get("撤回抗诉日期", ""))
            
            if protest_date and not withdraw_protest_date:
                return build_error_result(
                    row, "R049", "核对是否存在漏填撤回抗诉日期情况",
                    "核对是否漏填撤回抗诉日期，需人工审核",
                    提出抗诉日期=protest_date,
                    撤回抗诉日期=withdraw_protest_date,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R049",
            "规则名称": "核对是否存在漏填撤回抗诉日期情况",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R050(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            first_instance_sentence = clean(row.get("一审宣告刑", ""))
            
            if first_instance_sentence == "不负刑事责任":
                return build_error_result(
                    row, "R050", "判决结果为不负刑事责任的，需再核对",
                    "核对不负刑事责任案件，与45条一并核对",
                    一审宣告刑=first_instance_sentence,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R050",
            "规则名称": "判决结果为不负刑事责任的，需再核对",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R051(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_procedure = clean(row.get("庭审程序", ""))
            
            if trial_procedure != "速裁程序":
                return None
            
            identified_legal情节 = clean(row.get("审查认定的法定情节", ""))
            judged_legal情节 = clean(row.get("裁判认定法定情节", ""))
            case_confession = clean(row.get("全案_审查起诉阶段检察机关适用认罪认罚", ""))
            person_confession = clean(row.get("审查起诉阶段检察机关适用认罪认罚", ""))
            age_at_crime = clean(row.get("作案时年龄", ""))
            
            if "限定刑事责任能力人" in identified_legal情节 or "聋哑或盲人" in identified_legal情节 or \
               "限定刑事责任能力人" in judged_legal情节 or "聋哑或盲人" in judged_legal情节:
                return build_error_result(
                    row, "R051", "未成年人、盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用速裁程序",
                    "盲聋哑、限定行为能力人，核实是否错填速裁程序",
                    庭审程序=trial_procedure,
                    审查认定的法定情节=identified_legal情节,
                    裁判认定法定情节=judged_legal情节,
                    source_table=source_table
                )
            
            if case_confession == "是" and person_confession == "否":
                return build_error_result(
                    row, "R051", "未成年人、盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用速裁程序",
                    "共犯中部分被告人不认罪，核实是否错填速裁程序",
                    庭审程序=trial_procedure,
                    全案_审查起诉阶段检察机关适用认罪认罚=case_confession,
                    审查起诉阶段检察机关适用认罪认罚=person_confession,
                    source_table=source_table
                )
            
            if age_at_crime.isdigit() and int(age_at_crime) < 18:
                return build_error_result(
                    row, "R051", "未成年人、盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用速裁程序",
                    "未成年人，核实是否错填速裁程序",
                    庭审程序=trial_procedure,
                    作案时年龄=age_at_crime,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R051",
            "规则名称": "未成年人、盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用速裁程序",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R052(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_procedure = clean(row.get("庭审程序", ""))
            
            if trial_procedure != "简易程序":
                return None
            
            identified_legal情节 = clean(row.get("审查认定的法定情节", ""))
            judged_legal情节 = clean(row.get("裁判认定法定情节", ""))
            case_confession = clean(row.get("全案_审查起诉阶段检察机关适用认罪认罚", ""))
            person_confession = clean(row.get("审查起诉阶段检察机关适用认罪认罚", ""))
            
            if "限定刑事责任能力人" in identified_legal情节 or "聋哑或盲人" in identified_legal情节 or \
               "限定刑事责任能力人" in judged_legal情节 or "聋哑或盲人" in judged_legal情节:
                return build_error_result(
                    row, "R052", "盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用简易程序",
                    "盲聋哑、限定行为能力人，核实是否错填简易程序",
                    庭审程序=trial_procedure,
                    审查认定的法定情节=identified_legal情节,
                    裁判认定法定情节=judged_legal情节,
                    source_table=source_table
                )
            
            if case_confession == "是" and person_confession == "否":
                return build_error_result(
                    row, "R052", "盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用简易程序",
                    "共犯中部分被告人不认罪，核实是否错填简易程序",
                    庭审程序=trial_procedure,
                    全案_审查起诉阶段检察机关适用认罪认罚=case_confession,
                    审查起诉阶段检察机关适用认罪认罚=person_confession,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R052",
            "规则名称": "盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用简易程序",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R053(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_procedure = clean(row.get("庭审程序", ""))
            
            if trial_procedure != "速裁程序":
                return None
            
            first_instance_sentence_term = clean(row.get("一审宣告刑刑期", ""))
            first_instance_sentence = clean(row.get("一审宣告刑", ""))
            
            if (first_instance_sentence_term.isdigit() and int(first_instance_sentence_term) > 36) or \
               "无期" in first_instance_sentence:
                return build_error_result(
                    row, "R053", "可能判处有期徒刑三年以下刑罚的可以建议法院适用速裁程序（刑事诉讼规则第437条）",
                    "有期徒刑三年以上适用速裁程序，请核对",
                    庭审程序=trial_procedure,
                    一审宣告刑刑期=first_instance_sentence_term,
                    一审宣告刑=first_instance_sentence,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R053",
            "规则名称": "可能判处有期徒刑三年以下刑罚的可以建议法院适用速裁程序（刑事诉讼规则第437条）",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R054(row, case_person_summary, source_table="010101表", other_df=None):
    try:
        if source_table == "010101表" and other_df is not None:
            case_result = clean(row.get("审结处理结果", ""))
            if case_result != "批准逮捕":
                return None
            
            suspect_id = clean(row.get("嫌疑人编号", ""))
            if not suspect_id:
                return None
            
            mask = other_df["嫌疑人编号"].apply(clean) == suspect_id
            matches = other_df[mask]
            if not matches.empty:
                prosec_row = matches.iloc[0]
                审结处理情况 = clean(prosec_row.get("审结处理情况", ""))
                if 审结处理情况 != "起诉":
                    return None
                
                强制措施 = clean(prosec_row.get("审结前强制措施", ""))
                if 强制措施 != "逮捕":
                    return build_error_result(
                        row, "R054", "经审查逮捕程序中审结情况为批准逮捕，且强制措施尚未变更，但在审查起诉环节漏填“逮捕”的强制措施",
                        "批捕的嫌疑人，核实审结前强制措施是否错填",
                        审结处理结果=case_result,
                        嫌疑人编号=suspect_id,
                        审结处理情况=审结处理情况,
                        审结前强制措施=强制措施,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R054",
            "规则名称": "经审查逮捕程序中审结情况为批准逮捕，且强制措施尚未变更，但在审查起诉环节漏填“逮捕”的强制措施",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R055(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None
            
            current_date = datetime.now()
            
            first_delay_date = clean(row.get("一次建议法院延期审理日期", ""))
            first_recover_date = clean(row.get("一次建议法院恢复审理日期", ""))
            second_delay_date = clean(row.get("二次建议法院延期审理日期", ""))
            second_recover_date = clean(row.get("二次建议法院恢复审理日期", ""))
            
            parsed_first_delay = parse_date(first_delay_date)
            if parsed_first_delay:
                days_diff = (current_date - parsed_first_delay).days
                if days_diff > 31 and not first_recover_date:
                    return build_error_result(
                        row, "R055", "建议延期审理后，超过一个月未填录建议恢复审理日期",
                        "一次建议延期审理，未填录建议恢复审理日期",
                        一次建议法院延期审理日期=first_delay_date,
                        一次建议法院恢复审理日期=first_recover_date,
                        source_table=source_table
                    )
            
            parsed_second_delay = parse_date(second_delay_date)
            if parsed_second_delay:
                days_diff = (current_date - parsed_second_delay).days
                if days_diff > 32 and not second_recover_date:
                    return build_error_result(
                        row, "R055", "建议延期审理后，超过一个月未填录建议恢复审理日期",
                        "二次建议延期审理，超一个月未填录建议恢复审理日期",
                        二次建议法院延期审理日期=second_delay_date,
                        二次建议法院恢复审理日期=second_recover_date,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R055",
            "规则名称": "建议延期审理后，超过一个月未填录建议恢复审理日期",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R056(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            def parse_date(date_str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    try:
                        return datetime.strptime(date_str, "%Y/%m/%d")
                    except:
                        return None
            
            current_date = datetime.now()
            
            court_delay_date = clean(row.get("法院决定延期审理日期", ""))
            court_recover_date = clean(row.get("法院恢复审理日期", ""))
            
            parsed_court_delay = parse_date(court_delay_date)
            if parsed_court_delay:
                days_diff = (current_date - parsed_court_delay).days
                if days_diff > 94 and not court_recover_date:
                    return build_error_result(
                        row, "R056", "法院决定延期审后超过3个月未填录法院恢复审理日期",
                        "法院延期审理，超3个月未填录恢复审理日期",
                        法院决定延期审理日期=court_delay_date,
                        法院恢复审理日期=court_recover_date,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R056",
            "规则名称": "法院决定延期审后超过3个月未填录法院恢复审理日期",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R057(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_confession = clean(row.get("审判阶段适用认罪认罚情况", ""))
            prosecution_confession = clean(row.get("审查起诉阶段检察机关适用认罪认罚", ""))
            trial_apply = clean(row.get("审判阶段适用认罪认罚", ""))
            
            if "均适用" in trial_confession:
                if prosecution_confession == "否" or trial_apply == "否":
                    return build_error_result(
                        row, "R057", "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        审判阶段适用认罪认罚情况=trial_confession,
                        审查起诉阶段检察机关适用认罪认罚=prosecution_confession,
                        审判阶段适用认罪认罚=trial_apply,
                        source_table=source_table
                    )
            elif "均未适用" in trial_confession:
                if prosecution_confession == "是" or trial_apply == "是":
                    return build_error_result(
                        row, "R057", "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        审判阶段适用认罪认罚情况=trial_confession,
                        审查起诉阶段检察机关适用认罪认罚=prosecution_confession,
                        审判阶段适用认罪认罚=trial_apply,
                        source_table=source_table
                    )
            elif "审判阶段未适用" in trial_confession:
                if prosecution_confession == "否" or trial_apply == "是":
                    return build_error_result(
                        row, "R057", "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        审判阶段适用认罪认罚情况=trial_confession,
                        审查起诉阶段检察机关适用认罪认罚=prosecution_confession,
                        审判阶段适用认罪认罚=trial_apply,
                        source_table=source_table
                    )
            elif "审查起诉未适用" in trial_confession:
                if prosecution_confession == "是" or trial_apply == "否":
                    return build_error_result(
                        row, "R057", "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
                        审判阶段适用认罪认罚情况=trial_confession,
                        审查起诉阶段检察机关适用认罪认罚=prosecution_confession,
                        审判阶段适用认罪认罚=trial_apply,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R057",
            "规则名称": "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R058(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            court_accept = clean(row.get("是否法院采纳量刑建议", ""))
            prosecution_suggest = clean(row.get("审查起诉阶段提出量刑建议类型", ""))
            trial_suggest = clean(row.get("审判阶段提出（调整）量刑建议类型", ""))
            
            if court_accept == "是" and not prosecution_suggest and not trial_suggest:
                return build_error_result(
                    row, "R058", "法院采纳量刑建议填录“是”，审查起诉阶段量刑建议情况或者审判阶段提出（调整）量刑建议情况却漏填",
                    "法院采纳量刑建议填录“是”，漏填提出量刑建议情况",
                    是否法院采纳量刑建议=court_accept,
                    审查起诉阶段提出量刑建议类型=prosecution_suggest,
                    审判阶段提出_调整_量刑建议类型=trial_suggest,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R058",
            "规则名称": "法院采纳量刑建议填录“是”，审查起诉阶段量刑建议情况或者审判阶段提出（调整）量刑建议情况却漏填",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R059(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            is_unit_crime = clean(row.get("是否为单位", ""))
            first_instance_sentence = clean(row.get("一审宣告刑", ""))
            
            if is_unit_crime == "是" and ("有期徒刑" in first_instance_sentence or "拘役" in first_instance_sentence):
                return build_error_result(
                    row, "R059", "犯罪嫌疑人“是否为单位”填录(是),“一审判决情况”填录（拘役）或者（有期徒刑）",
                    "单位犯罪判处拘役和有期徒刑，请核对",
                    是否单位犯罪=is_unit_crime,
                    一审宣告刑=first_instance_sentence,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R059",
            "规则名称": "犯罪嫌疑人“是否为单位”填录(是),“一审判决情况”填录（拘役）或者（有期徒刑）",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R060(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            trial_procedure = clean(row.get("庭审程序", ""))
            convert_to_simple_date = clean(row.get("速裁程序转为简易程序日期", ""))
            convert_to_normal_date = clean(row.get("速裁程序转为普通程序日期", ""))
            court_date = clean(row.get("出庭日期", ""))
            first_instance_judgment_date = clean(row.get("一审判决日期", ""))
            
            if trial_procedure == "速裁程序" and not convert_to_simple_date and not convert_to_normal_date and court_date != first_instance_judgment_date:
                return build_error_result(
                    row, "R060", "速裁程序审结的案件，出庭日期和一审判决日期应为同一日",
                    "速裁程序案件，出庭日期和一审判决日期不为同一天，请核对",
                    庭审程序=trial_procedure,
                    出庭日期=court_date,
                    一审判决日期=first_instance_judgment_date,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R060",
            "规则名称": "速裁程序审结的案件，出庭日期和一审判决日期应为同一日",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R061(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            close_case_result = clean(row.get("审结处理情况", ""))
            first_instance_judgment_date = clean(row.get("一审判决日期", ""))
            if close_case_result != "起诉" or not first_instance_judgment_date:
                return None
            
            close_case_reason = clean(row.get("审结案由", ""))
            close_case_other_reasons = clean(row.get("审结其他案由", ""))
            judgment_reason = clean(row.get("一审判决罪名/案由", ""))
            judgment_other_reasons = clean(row.get("一审判决其他罪名/案由", ""))
            change_prosecution = clean(row.get("裁判改变起诉情形", ""))
            
            close_reasons = set()
            if close_case_reason:
                close_reasons.add(close_case_reason.strip())
            if close_case_other_reasons:
                for reason in close_case_other_reasons.split(","):
                    reason = reason.strip()
                    if reason:
                        close_reasons.add(reason)
            
            judgment_reasons = set()
            if judgment_reason:
                judgment_reasons.add(judgment_reason.strip())
            if judgment_other_reasons:
                for reason in judgment_other_reasons.split(","):
                    reason = reason.strip()
                    if reason:
                        judgment_reasons.add(reason)
            
            has_change_qualification = "改变定性" in change_prosecution
            
            if close_reasons == judgment_reasons:
                if has_change_qualification:
                    return build_error_result(
                        row, "R061", "审结罪名和判决罪名不一致时，裁判改变起诉情况应当填录“改变定性”",
                        "个人审结罪名与裁判罪名一致，核查错填改变定性",
                        审结案由=close_case_reason,
                        审结其他案由=close_case_other_reasons,
                        一审判决罪名=judgment_reason,
                        一审判决其他罪名=judgment_other_reasons,
                        裁判改变起诉情形=change_prosecution,
                        source_table=source_table
                    )
            else:
                if not has_change_qualification:
                    return build_error_result(
                        row, "R061", "审结罪名和判决罪名不一致时，裁判改变起诉情况应当填录“改变定性”",
                        "个人审结罪名与裁判罪名不一致，核查漏填改变定性",
                        审结案由=close_case_reason,
                        审结其他案由=close_case_other_reasons,
                        一审判决罪名=judgment_reason,
                        一审判决其他罪名=judgment_other_reasons,
                        裁判改变起诉情形=change_prosecution,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R061",
            "规则名称": "审结罪名和判决罪名不一致时，裁判改变起诉情况应当填录“改变定性”",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R062(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            protest_date = clean(row.get("提出抗诉日期", ""))
            protest_reason = clean(row.get("提出抗诉理由", ""))
            
            if protest_date and not protest_reason:
                return build_error_result(
                    row, "R062", "漏填抗诉理由",
                    "抗诉案件，漏填提出抗诉理由",
                    提出抗诉日期=protest_date,
                    提出抗诉理由=protest_reason,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R062",
            "规则名称": "漏填抗诉理由",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R063(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            protest_date = clean(row.get("提出抗诉日期", ""))
            adopt_protest = clean(row.get("是否采纳抗诉意见", ""))
            
            if protest_date and not adopt_protest:
                return build_error_result(
                    row, "R063", "漏填采纳抗诉意见",
                    "抗诉案件，漏填采纳抗诉意见",
                    提出抗诉日期=protest_date,
                    采纳抗诉意见=adopt_protest,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R063",
            "规则名称": "漏填采纳抗诉意见",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R064(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            protest_date = clean(row.get("提出抗诉日期", ""))
            higher_procuratorate_result = clean(row.get("上级检察院审查结果", ""))
            
            if protest_date and not higher_procuratorate_result:
                return build_error_result(
                    row, "R064", "提出抗诉后漏填上级检察院审查结果",
                    "抗诉案件，漏填上级检察院审查结果",
                    提出抗诉日期=protest_date,
                    上级检察院审查结果=higher_procuratorate_result,
                    source_table=source_table
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R064",
            "规则名称": "提出抗诉后漏填上级检察院审查结果",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R066(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            ajbh = clean(row.get("部门受案号"))
            qkaj_rzrf = clean(row.get("全案_审查起诉阶段检察机关适用认罪认罚", ""))
            gr_rzrf = clean(row.get("审查起诉阶段检察机关适用认罪认罚", ""))
            
            person_count = case_person_summary.get(ajbh, {"count": 1})["count"]
            
            if person_count == 1:
                if qkaj_rzrf != gr_rzrf:
                    return build_error_result(
                        row, "R066", "检察机关适用认罪认罚人卡与全案矛盾",
                        "一案一人，个人与全案适用认罪认罚不一致",
                        全案_审查起诉阶段检察机关适用认罪认罚=qkaj_rzrf,
                        审查起诉阶段检察机关适用认罪认罚=gr_rzrf,
                        source_table=source_table
                    )
            else:
                if qkaj_rzrf == "否" and gr_rzrf == "是":
                    return build_error_result(
                        row, "R066", "检察机关适用认罪认罚人卡与全案矛盾",
                        "一案多人，一人认罪认罚，全案适用认罪认罚",
                        全案_审查起诉阶段检察机关适用认罪认罚=qkaj_rzrf,
                        审查起诉阶段检察机关适用认罪认罚=gr_rzrf,
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R066",
            "规则名称": "检察机关适用认罪认罚人卡与全案矛盾",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R067(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            ljyj_type = clean(row.get("审查起诉阶段提出量刑建议类型", ""))
            ljyj_xz = clean(row.get("审查起诉阶段量刑建议刑种", ""))
            ljyj_yq_xq = clean(row.get("审查起诉阶段量刑建议（自）有期徒刑刑期", ""))
            ljyj_jy_xq = clean(row.get("审查起诉阶段量刑建议（自）拘役刑期", ""))
            ljyj_gz_xq = clean(row.get("审查起诉阶段量刑建议（自）管制刑期", ""))
            
            if ljyj_type:
                if not ljyj_xz:
                    return build_error_result(
                        row, "R067", "填报提出量刑建议，但漏填具体刑种、刑期",
                        "提出量刑建议，漏填量刑建议刑种",
                        审查起诉阶段提出量刑建议类型=ljyj_type,
                        审查起诉阶段量刑建议刑种=ljyj_xz,
                        source_table=source_table
                    )
                if "有期" in ljyj_xz and not ljyj_yq_xq:
                    return build_error_result(
                        row, "R067", "填报提出量刑建议，但漏填具体刑种、刑期",
                        "提出量刑建议，漏填量刑建议刑期",
                        审查起诉阶段提出量刑建议类型=ljyj_type,
                        审查起诉阶段量刑建议刑种=ljyj_xz,
                        **{"审查起诉阶段量刑建议（自）有期徒刑刑期": ljyj_yq_xq},
                        source_table=source_table
                    )
                if "拘役" in ljyj_xz and not ljyj_jy_xq:
                    return build_error_result(
                        row, "R067", "填报提出量刑建议，但漏填具体刑种、刑期",
                        "提出量刑建议，漏填量刑建议刑期",
                        审查起诉阶段提出量刑建议类型=ljyj_type,
                        审查起诉阶段量刑建议刑种=ljyj_xz,
                        **{"审查起诉阶段量刑建议（自）拘役刑期": ljyj_jy_xq},
                        source_table=source_table
                    )
                if "管制" in ljyj_xz and not ljyj_gz_xq:
                    return build_error_result(
                        row, "R067", "填报提出量刑建议，但漏填具体刑种、刑期",
                        "提出量刑建议，漏填量刑建议刑期",
                        审查起诉阶段提出量刑建议类型=ljyj_type,
                        审查起诉阶段量刑建议刑种=ljyj_xz,
                        **{"审查起诉阶段量刑建议（自）管制刑期": ljyj_gz_xq},
                        source_table=source_table
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R067",
            "规则名称": "填报提出量刑建议，但漏填具体刑种、刑期",
            "冲突说明": "校验异常：字段缺失"
        }

def check_R068(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            sjqs_ljyj_z_fj = clean(row.get("审查起诉阶段量刑建议（自）罚金数额", ""))
            sjqs_ljyj_d_fj = clean(row.get("审查起诉阶段量刑建议（到）罚金数额", ""))
            spjd_ljyj_z_fj = clean(row.get("审判阶段量刑建议（自）罚金数额", ""))
            spjd_ljyj_d_fj = clean(row.get("审判阶段量刑建议（到）罚金数额", ""))
            
            conflict_fields = {}
            has_conflict = False
            
            def check_field(value, field_name):
                nonlocal has_conflict
                if value:
                    try:
                        num_value = float(value)
                        if num_value > 499:
                            has_conflict = True
                            conflict_fields[field_name] = value
                    except:
                        pass
            
            check_field(sjqs_ljyj_z_fj, "审查起诉阶段量刑建议（自）罚金数额")
            check_field(sjqs_ljyj_d_fj, "审查起诉阶段量刑建议（到）罚金数额")
            check_field(spjd_ljyj_z_fj, "审判阶段量刑建议（自）罚金数额")
            check_field(spjd_ljyj_d_fj, "审判阶段量刑建议（到）罚金数额")
            
            if has_conflict:
                return build_error_result(
                    row, "R068", '量刑罚金单位"万元"',
                    "核对量刑罚金，单位为万元",
                    **{"审查起诉阶段量刑建议（自）罚金数额": sjqs_ljyj_z_fj},
                    **{"审查起诉阶段量刑建议（到）罚金数额": sjqs_ljyj_d_fj},
                    **{"审判阶段量刑建议（自）罚金数额": spjd_ljyj_z_fj},
                    **{"审判阶段量刑建议（到）罚金数额": spjd_ljyj_d_fj},
                    source_table=source_table,
                    #
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R068",
            "规则名称": '量刑罚金单位"万元"',
            "冲突说明": "校验异常：字段缺失"
        }

def check_R069(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            ljyj_type = clean(row.get("审查起诉阶段提出量刑建议类型"))
            ljyj_z_yqsx = row.get("审查起诉阶段量刑建议（自）有期徒刑刑期")
            ljyj_d_yqsx = row.get("审查起诉阶段量刑建议（到）有期徒刑刑期")
            
            if ljyj_type == "幅度刑量刑建议":
                try:
                    d_value = float(ljyj_d_yqsx) if pd.notna(ljyj_d_yqsx) else 0
                    z_value = float(ljyj_z_yqsx) if pd.notna(ljyj_z_yqsx) else 0
                    
                    if d_value * 0.75 > z_value:
                        return build_error_result(
                            row, "R069", "幅度刑量刑建议幅度过于宽泛（超过上限值的25%，可自定义）",
                            "量刑建议幅度过于宽泛（超过上限值的25%，可自定义）",
                            source_table=source_table,
                            审查起诉阶段提出量刑建议类型=ljyj_type,
                            **{"审查起诉阶段量刑建议（自）有期徒刑刑期": ljyj_z_yqsx},
                            **{"审查起诉阶段量刑建议（到）有期徒刑刑期": ljyj_d_yqsx}
                        )
                except:
                    pass
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R069",
            "规则名称": "幅度刑量刑建议幅度过于宽泛（超过上限值的25%，可自定义）",
            "冲突说明": "校验异常：字段缺失"
        }


def check_R070(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            hongkong_macao_taiwan = clean(row.get("涉港澳台侨犯罪案件", ""))
            foreign = clean(row.get("是否涉外案件", ""))
            political_security = clean(row.get("涉政治安全类案件", ""))
            if "涉" in hongkong_macao_taiwan or foreign == "是" or political_security == "是":
                return build_error_result(
                    row, "R070", "起诉核对涉外、涉侨、涉台、涉港澳案件及涉及政治安全",
                    "涉港澳台侨、涉外或政治安全案件，请核对",
                    source_table=source_table,
                    审结案由=clean(row.get("审结案由")),
                    审结其他案由=clean(row.get("审结其他案由")),
                    涉港澳台侨犯罪案件=hongkong_macao_taiwan,
                    是否涉外案件=foreign,
                    涉政治安全类案件=political_security
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R070",
            "规则名称": "起诉核对涉外、涉侨、涉台、涉港澳案件及涉及政治安全",
            "冲突说明": "校验异常：字段缺失",
            "来源表": source_table
        }


def check_R071(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            non_public_econ = clean(row.get("是否影响非公有制经济发展案件", ""))
            if non_public_econ == "是":
                return build_error_result(
                    row, "R071", "起诉案卡是否影响非公经济核对",
                    "是否影响非公有制经济发展案件为是，请核对审结案由",
                    source_table=source_table,
                    审结案由=clean(row.get("审结案由")),
                    审结其他案由=clean(row.get("审结其他案由")),
                    是否影响非公有制经济发展案件=non_public_econ,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R071",
            "规则名称": "起诉案卡是否影响非公经济核对",
            "冲突说明": "校验异常：字段缺失",
            "来源表": source_table
        }


def check_R072(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            identified_legal情节 = clean(row.get("审查认定的法定情节", ""))
            criminal_record = clean(row.get("前科", ""))
            
            if "累犯" in identified_legal情节 and "刑事处罚" not in criminal_record:
                return build_error_result(
                    row, "R072", "审查起诉案件，审结认定累犯，但前科填“无”",
                    "审结认定累犯，请核对前科是否漏填",
                    source_table=source_table,
                    审查认定的法定情节=identified_legal情节,
                    前科=criminal_record,
                )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R072",
            "规则名称": "审查起诉案件，审结认定累犯，但前科填“无”",
            "冲突说明": "校验异常：字段缺失",
            "来源表": source_table
        }


def check_R076(row, case_person_summary, source_table="020101表", other_df=None):
    try:
        if source_table == "020101表":
            ip_crimes = [
                "假冒注册商标罪",
                "销售假冒注册商标的商品罪",
                "非法制造、销售非法制造的注册商标标识罪",
                "假冒专利罪",
                "侵犯著作权罪",
                "销售侵权复制品罪",
                "侵犯商业秘密罪",
                "为境外窃取、刺探、收买、非法提供商业秘密罪"
            ]
            
            case_charge = clean(row.get("审结案由", ""))
            other_charge = clean(row.get("审结其他案由", ""))
            
            has_ip_crime = any(crime in case_charge or crime in other_charge for crime in ip_crimes)
            
            if has_ip_crime:
                enterprise_type = clean(row.get("受侵犯企业类型", ""))
                fund_source = clean(row.get("受侵犯企业资金来源地", ""))
                country_region = clean(row.get("受侵犯企业所属国别地区", ""))
                
                if not enterprise_type or not fund_source or not country_region:
                    return build_error_result(
                        row, "R076", "知识产权类案件中漏填受侵害情况",
                        "知识产权类案件，请核对是否漏填受侵害情况",
                        source_table=source_table,
                        审结案由=case_charge,
                        审结其他案由=other_charge,
                        受侵犯企业类型=enterprise_type,
                        受侵犯企业资金来源地=fund_source,
                        受侵犯企业所属国别地区=country_region,
                    )
        return None
    except:
        return {
            "部门受案号": clean(row.get("部门受案号")),
            "规则编号": "R076",
            "规则名称": "知识产权类案件中漏填受侵害情况",
            "冲突说明": "校验异常：字段缺失",
            "来源表": source_table
        }

# ====================== 规则引擎注册 =====================
rule_list = [
    ("R001", "审查逮捕案件受理时错填移送罪名", check_R001),
    ("R002", "是否存在其他不捕情形", check_R002),
    ("R003", '审结结果为批准逮捕案件，应当逮捕情形含"曾经故意犯罪"，前科填"无"', check_R003),
    ("R004", "受理案由为普通刑事犯罪案由，但侦查机关或移送单位错填为检察机关", check_R004),
    ("R005", "逮捕案卡全案审结结果与个人审结结果矛盾", check_R005),
    ("R006", "审查逮捕案件审结日期错填", check_R006),
    ("R007", "审查逮捕审结后长期未填录执行情况", check_R007),
    ("R008", "侵犯公民人身权利、民主权利及财产类案件中，“被害人情况”不应填录（无）", check_R008),
    ("R009", "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况", check_R009),
    ("R010", "审查逮捕案件审结案由为“掩犯罪所得、犯罪所得收益罪”时，漏填掩饰隐瞒犯罪所得具体情形", check_R010),
    ("R011", "核查审查逮捕、审查起诉案件特殊罪名", check_R011),
    ("R012", "核对审查逮捕案件，单人单案，审结案由为危险驾驶案件被批捕情况", check_R012),
    ("R014", "危险驾驶、交通肇事类案件填录为“涉及利用电信网络实施犯罪”", check_R014),
    ("R015", "逮捕、起诉案卡是否涉家庭暴力核对", check_R015),
    ("R016", "逮捕、起诉案卡是否涉校园暴力核对", check_R016),
    ("R017", "逮捕、起诉案卡涉医核对", check_R017),
    ("R018", "逮捕核对涉外、涉侨、涉台、涉港澳案件及涉及政治安全", check_R018),
    ("R019", "逮捕案卡是否影响非公经济核对", check_R019),
    ("R022", "在审查逮捕阶段填报为纠正的犯罪嫌疑人，在审查起诉阶段未填为“追捕对象”", check_R022),
    ("R023", "审查起诉阶段填报为“追捕对象”，审查逮捕阶段未填为纠正漏捕的嫌疑人", check_R023),
    ("R024", "核查纠正漏捕后续处理情况是否正确", check_R024),
    ("R025", "黑势力犯罪情况", check_R025),
    ("R026", "恶势力犯罪情况", check_R026),
    ("R027", "审查逮捕、审查起诉错填涉农资犯罪、涉食药品犯罪", check_R027),
    ("R028", "非未成年人附条件不起诉，或附条件不起诉罪名是否为四、五、六章犯罪", check_R028),
    ("R029", "一审公诉案件受理时错填移送罪名", check_R029),
    ("R030", "受理移送不起诉是否错填", check_R030),
    ("R031", "经济类、侵财类犯罪核对'涉及追赃挽损情况'", check_R031),
    ("R032", "侵犯公民人身权利、民主权利及财产类案件中，'被害人情况'不应填录（无）", check_R032),
    ("R033", "故意伤害、故意杀人、过失致人重伤死亡、交通肇事等应该有案件造成伤亡情况的罪名漏填案件造成伤亡情况", check_R033),
    ("R034", "一审公诉案件受理后超过一年未审结", check_R034),
    ("R035", "一审公诉案件审结后超过一年无判决", check_R035),
    ("R036", "速裁程序办案时限十天，可能判处有期徒刑超过一年的，可以延长十五日", check_R036),
    ("R037", "同一犯罪嫌疑人既立案监督又纠正漏捕又纠正漏诉", check_R037),
    ("R038", "一审公诉案件退查后长期未重报且未填报“退查未重报”", check_R038),
    ("R039", "核对是否存在存疑不诉、绝对不诉案件适用认罪认罚", check_R039),
    ("R040", "核对相对不诉但未适用认罪认罚", check_R040),
    ("R041", "是否他院受理后改变管辖漏填", check_R041),
    ("R042", "核对一审公诉案件审结、追赃挽损填录异常问题", check_R042),
    ("R043", "撤回起诉后一个月案件无处理结果", check_R043),
    ("R044", "审查起诉改变侦(调)查认定的情形漏填", check_R044),
    ("R045", "核对无罪、撤回起诉、不负刑事责任案件填报情况，与50条一并核对。", check_R045),
    ("R046", "漏填提起公诉日期", check_R046),
    ("R047", "一审判决后未上诉抗诉长期不填报生效日期", check_R047),
    ("R048", "一审公诉填报案件裁判结果，人员裁判结果未填报", check_R048),
    ("R049", "核对是否存在漏填撤回抗诉日期情况", check_R049),
    ("R050", "判决结果为不负刑事责任的，需再核对", check_R050),
    ("R051", "未成年人、盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用速裁程序", check_R051),
    ("R052", "盲聋哑、限定行为能力人、共犯中部分被告人不认罪等不得适用简易程序", check_R052),
    ("R053", "可能判处有期徒刑三年以下刑罚的可以建议法院适用速裁程序（刑事诉讼规则第437条）", check_R053),
    ("R054", "经审查逮捕程序中审结情况为批准逮捕，且强制措施尚未变更，但在审查起诉环节漏填“逮捕”的强制措施", check_R054),
    ("R055", "建议延期审理后，超过一个月未填录建议恢复审理日期", check_R055),
    ("R056", "法院决定延期审后超过3个月未填录法院恢复审理日期", check_R056),
    ("R057", "审查起诉阶段适用认罪认罚情况与审判阶段适用认罪认罚情况存在矛盾", check_R057),
    ("R058", "法院采纳量刑建议填录“是”，审查起诉阶段量刑建议情况或者审判阶段提出（调整）量刑建议情况却漏填", check_R058),
    ("R059", "犯罪嫌疑人“是否为单位”填录(是),“一审判决情况”填录（拘役）或者（有期徒刑）", check_R059),
    ("R060", "速裁程序审结的案件，出庭日期和一审判决日期应为同一日", check_R060),
    ("R061", "审结罪名和判决罪名不一致时，裁判改变起诉情况应当填录“改变定性”", check_R061),
    ("R062", "漏填抗诉理由", check_R062),
    ("R063", "漏填采纳抗诉意见", check_R063),
    ("R064", "提出抗诉后漏填上级检察院审查结果", check_R064),
    ("R066", "检察机关适用认罪认罚人卡与全案矛盾", check_R066),
    ("R067", "填报提出量刑建议，但漏填具体刑种、刑期", check_R067),
    ("R068", '量刑罚金单位"万元"', check_R068),
    ("R069", "幅度刑量刑建议幅度过于宽泛（超过上限值的25%，可自定义）", check_R069),
    ("R070", "起诉核对涉外、涉侨、涉台、涉港澳案件及涉及政治安全", check_R070),
    ("R071", "起诉案卡是否影响非公经济核对", check_R071),
    ("R072", "审查起诉案件，审结认定累犯，但前科填“无”", check_R072),
    ("R076", "知识产权类案件中漏填受侵害情况", check_R076),
]
rule_checks = {}
last_selected_index = -1


# ====================== 规则适用表配置 =====================
RULE_APPLIES = {
    "R001": "010101表",
    "R002": "010101表",
    "R003": "010101表",
    "R004": "010101表",
    "R005": "010101表",
    "R006": "010101表",
    "R007": "010101表",
    "R008": "010101表",
    "R009": "010101表",
    "R010": "010101表",
    "R011": "both",
    "R012": "010101表",
    "R014": "both",
    "R015": "both",
    "R016": "both",
    "R017": "both",
    "R018": "010101表",
    "R019": "010101表",
    "R022": "010101表",
    "R023": "020101表",
    "R024": "010101表",
    "R025": "both",
    "R026": "both",
    "R027": "both",
    "R028": "020101表",
    "R029": "020101表",
    "R030": "020101表",
    "R031": "020101表",
    "R032": "020101表",
    "R033": "020101表",
    "R034": "020101表",
    "R035": "020101表",
    "R036": "020101表",
    "R037": "020101表",
    "R038": "020101表",
    "R039": "020101表",
    "R040": "020101表",
    "R041": "020101表",
    "R042": "020101表",
    "R043": "020101表",
    "R044": "020101表",
    "R045": "020101表",
    "R046": "020101表",
    "R047": "020101表",
    "R048": "020101表",
    "R049": "020101表",
    "R050": "020101表",
    "R051": "020101表",
    "R052": "020101表",
    "R053": "020101表",
    "R054": "010101表",
    "R055": "020101表",
    "R056": "020101表",
    "R057": "020101表",
    "R058": "020101表",
    "R059": "020101表",
    "R060": "020101表",
    "R061": "020101表",
    "R062": "020101表",
    "R063": "020101表",
    "R064": "020101表",
    "R066": "020101表",
    "R067": "020101表",
    "R068": "020101表",
    "R069": "020101表",
    "R070": "020101表",
    "R071": "020101表",
    "R072": "020101表",
    "R076": "020101表",
}

def run_all_rules(row, selected_codes, case_person_summary, source_table="010101表", other_df=None):
    errors = []
    for code, name, func in rule_list:
        if code in selected_codes and func is not None:
            applies_to = RULE_APPLIES.get(code, "010101表")
            if applies_to == "both" or applies_to == source_table:
                result = func(row, case_person_summary, source_table, other_df)
                if result:
                    errors.append(result)
    return errors


# ====================== 工具函数 ======================
def get_auth_status():
    try:
        status, days, expire = check_license()
        if status in ["normal", "warn"]:
            return True, days, expire
        elif status == "expired":
            return False, 0, "已过期"
        elif status == "inactive":
            return False, 0, "未激活"
        else:
            return False, 0, "授权异常"
    except:
        return False, 0, "授权异常"

def refresh_auth_label():
    if label_auth_info:
        ok, days, expire = get_auth_status()
        if ok:
            if days <= 3:
                color = "#fbbf24"
                bg_color = "#7c2d12"
                status_text = f"⚠️ 剩余 {days} 天 | {expire}"
            else:
                color = "#a7f3d0"
                bg_color = "#1e40af"
                status_text = f"✓ 剩余 {days} 天 | {expire}"
        else:
            color = "#fecaca"
            bg_color = "#7f1d1d"
            status_text = "✗ 授权失效"
        label_auth_info.config(text=status_text, fg=color, bg=bg_color)

def auto_scan_file():
    global path_arrest, path_prosec, path_filter
    if os.path.exists(FILE_ARREST): path_arrest = FILE_ARREST
    if os.path.exists(FILE_PROSEC): path_prosec = FILE_PROSEC
    if os.path.exists(FILE_FILTER): path_filter = FILE_FILTER

def get_skip_data():
    skip = set()
    if os.path.exists(FILE_FILTER):
        try:
            df = read_file(FILE_FILTER, dtype=str)
            for _, r in df.iterrows():
                xyr = clean(r.get("嫌疑人编号"))
                code = clean(r.get("规则编号"))
                if xyr and code: skip.add((xyr, code))
        except:
            pass
    return skip

# ====================== 规则操作 ======================
def select_all_rules():
    for v in rule_checks.values(): v.set(True)

def invert_select_rules():
    for v in rule_checks.values(): v.set(not v.get())

def clear_select_rules():
    for v in rule_checks.values(): v.set(False)

def on_rule_click(e, idx):
    global last_selected_index
    if e.state & 0x0001:
        if last_selected_index != -1:
            s = min(last_selected_index, idx)
            e_idx = max(last_selected_index, idx)
            for i in range(s, e_idx + 1):
                list(rule_checks.values())[i].set(True)
    last_selected_index = idx

# ====================== 文件操作 ======================
def read_file(path, dtype=None):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path, dtype=dtype, encoding='utf-8-sig')
    else:
        return pd.read_excel(path, dtype=dtype)

def upload_arrest():
    global path_arrest
    f = filedialog.askopenfilename(filetypes=[("数据文件", "*.xlsx *.xls *.csv")])
    if f:
        shutil.copy(f, FILE_ARREST)
        path_arrest = FILE_ARREST
        label_arrest.config(text="✓", fg="#059669")
        update_file_status()

def upload_prosec():
    global path_prosec
    f = filedialog.askopenfilename(filetypes=[("数据文件", "*.xlsx *.xls *.csv")])
    if f:
        shutil.copy(f, FILE_PROSEC)
        path_prosec = FILE_PROSEC
        label_prosec.config(text="✓", fg="#059669")
        update_file_status()

def upload_filter():
    global path_filter
    f = filedialog.askopenfilename(filetypes=[("数据文件", "*.xlsx *.xls *.csv")])
    if f:
        shutil.copy(f, FILE_FILTER)
        path_filter = FILE_FILTER
        label_filter.config(text="✓", fg="#059669")
        update_file_status()

def update_file_status():
    global label_arrest, label_prosec, label_filter
    global label_arrest_count, label_prosec_count, label_filter_count
    global label_arrest_tip, label_prosec_tip, label_filter_tip
    if not status_vars:
        return
    try:
        df1_count = len(read_file(path_arrest)) if path_arrest and os.path.exists(path_arrest) else 0
    except:
        df1_count = 0
    try:
        df2_count = len(read_file(path_prosec)) if path_prosec and os.path.exists(path_prosec) else 0
    except:
        df2_count = 0
    try:
        if path_filter and os.path.exists(path_filter):
            filter_count = len(read_file(path_filter))
        else:
            filter_count = 0
    except:
        filter_count = 0
    
    # 更新文件栏的状态显示
    if label_arrest:
        if path_arrest and os.path.exists(path_arrest):
            label_arrest.config(text="✓", fg="#059669")
            if label_arrest_count:
                label_arrest_count.config(text=f"{df1_count}条")
            if label_arrest_tip:
                label_arrest_tip.config(text="已自动读取文件", fg="#059669")
        else:
            label_arrest.config(text="✗", fg="#ef4444")
            if label_arrest_count:
                label_arrest_count.config(text="")
            if label_arrest_tip:
                label_arrest_tip.config(text="未上传，请上传数据", fg="#ef4444")
    
    if label_prosec:
        if path_prosec and os.path.exists(path_prosec):
            label_prosec.config(text="✓", fg="#059669")
            if label_prosec_count:
                label_prosec_count.config(text=f"{df2_count}条")
            if label_prosec_tip:
                label_prosec_tip.config(text="已自动读取文件", fg="#059669")
        else:
            label_prosec.config(text="✗", fg="#ef4444")
            if label_prosec_count:
                label_prosec_count.config(text="")
            if label_prosec_tip:
                label_prosec_tip.config(text="未上传，请上传数据", fg="#ef4444")
    
    if label_filter:
        if path_filter and os.path.exists(path_filter):
            label_filter.config(text="✓", fg="#059669")
            if label_filter_count:
                label_filter_count.config(text=f"{filter_count}条")
            if label_filter_tip:
                label_filter_tip.config(text="已自动读取文件", fg="#059669")
        else:
            label_filter.config(text="✗", fg="#ef4444")
            if label_filter_count:
                label_filter_count.config(text="")
            if label_filter_tip:
                label_filter_tip.config(text="未上传，请上传数据", fg="#ef4444")
    
    # 更新状态信息栏
    try:
        if status_vars and "info1" in status_vars:
            status_vars["info1"].set(f"{df1_count} 条" if df1_count > 0 else "--")
        if status_vars and "info2" in status_vars:
            status_vars["info2"].set(f"{df2_count} 条" if df2_count > 0 else "--")
        if status_vars and "info3" in status_vars:
            status_vars["info3"].set(f"{filter_count} 条" if filter_count > 0 else "--")
        if status_vars and "result" in status_vars:
            status_vars["result"].set("等待校验")
    except:
        pass

# ====================== 主校验程序 ======================
def start_check():
    try:
        if not is_license_valid():
            show_custom_dialog("授权失效", "请先激活授权后再使用！", "error")
            return
        
        missing = []
        if not path_arrest or not os.path.exists(path_arrest):
            missing.append("010101表")
        if not path_prosec or not os.path.exists(path_prosec):
            missing.append("020101表")
        if not path_filter or not os.path.exists(path_filter):
            missing.append("已核对清单")
        
        if missing:
            if len(missing) == 1:
                msg = f"请上传{missing[0]}"
            elif len(missing) == 2:
                msg = f"请上传{missing[0]}和{missing[1]}"
            else:
                msg = f"请上传{missing[0]}、{missing[1]}和{missing[2]}"
            show_custom_dialog("需要上传文件", msg, "warning")
            return
        
        selected = [k for k, v in rule_checks.items() if v.get()]
        if not selected:
            show_custom_dialog("请选择规则", "请至少选择一条规则来进行校验", "warning")
            return
        
        try:
            if status_vars and "result" in status_vars:
                status_vars["result"].set("校验中...")
            if status_vars and "conflicts" in status_vars:
                status_vars["conflicts"].set("--")
            if status_vars and "status" in status_vars:
                status_vars["status"].set("🔄 正在校验中，请稍候...")
            if status_label:
                status_label.config(bg="#f59e0b", fg="#ffffff")
        except:
            pass
        
        skip_set = get_skip_data()
        df1 = read_file(path_arrest, dtype=str)
        df2 = read_file(path_prosec, dtype=str)
        
        case_person_summary = {}
        for _, row in df1.iterrows():
            ajbh = clean(row["部门受案号"])
            if ajbh not in case_person_summary:
                case_person_summary[ajbh] = {"count": 0, "total_set": set()}
            case_person_summary[ajbh]["count"] += 1
            sx = parse_charges(clean(row["涉嫌案由"]))
            sxt = parse_charges(clean(row["涉嫌其他案由"]))
            case_person_summary[ajbh]["total_set"] |= sx | sxt

        case_person_summary_prosec = {}
        xyrbh_count = {}
        for _, row in df2.iterrows():
            ajbh = clean(row["部门受案号"])
            xyrbh = clean(row.get("嫌疑人编号", ""))
            
            if ajbh not in case_person_summary_prosec:
                case_person_summary_prosec[ajbh] = {"count": 0, "total_set": set()}
            case_person_summary_prosec[ajbh]["count"] += 1
            sx = parse_charges(clean(row.get("移诉案由", "")))
            sxt = parse_charges(clean(row.get("移诉其他案由", "")))
            case_person_summary_prosec[ajbh]["total_set"] |= sx | sxt
            
            if xyrbh:
                xyrbh_count[xyrbh] = xyrbh_count.get(xyrbh, 0) + 1
        
        case_person_summary_prosec["xyrbh_count"] = xyrbh_count
        
        df1_count = len(df1)
        df2_count = len(df2)
        filter_count = len(skip_set)
        
        try:
            if status_vars and "info1" in status_vars:
                status_vars["info1"].set(f"{df1_count} 条")
            if status_vars and "info2" in status_vars:
                status_vars["info2"].set(f"{df2_count} 条")
            if status_vars and "info3" in status_vars:
                status_vars["info3"].set(f"{filter_count} 条")
        except:
            pass
        
        root.update()
        all_err = []

        selected_table = "both"

        if selected_table in ["010101表", "both"]:
            for i, row in df1.iterrows():
                progress = f"{i + 1}/{df1_count} ({int((i + 1) / df1_count * 100)}%)"
                try:
                    if status_vars and "progress" in status_vars:
                        status_vars["progress"].set(progress)
                except:
                    pass
                root.update()

                other_df = df2 if selected_table == "both" else None
                errs = run_all_rules(row, selected, case_person_summary, source_table="010101表", other_df=other_df)
                for e in errs:
                    xyr = e.get("嫌疑人编号", "")
                    code = e.get("规则编号", "")
                    if (xyr, code) not in skip_set:
                        all_err.append(e)

        if selected_table in ["020101表", "both"]:
            for i, row in df2.iterrows():
                progress = f"020101表 {i + 1}/{df2_count} ({int((i + 1) / df2_count * 100)}%)"
                try:
                    if status_vars and "progress" in status_vars:
                        status_vars["progress"].set(progress)
                except:
                    pass
                root.update()

                other_df = df1 if selected_table == "both" else None
                errs = run_all_rules(row, selected, case_person_summary_prosec, source_table="020101表", other_df=other_df)
                for e in errs:
                    xyr = e.get("嫌疑人编号", "")
                    code = e.get("规则编号", "")
                    if (xyr, code) not in skip_set:
                        all_err.append(e)

        if all_err:
            os.makedirs("审核结果", exist_ok=True)
            fn = f"审核结果/案卡冲突_{datetime.now():%Y%m%d%H%M%S}.xlsx"

            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from collections import defaultdict

            wb = Workbook()
            ws = wb.active
            ws.title = "冲突清单"

            rule_group = defaultdict(list)
            for e in all_err:
                code = e.get("规则编号", "")
                rule_group[code].append(e)

            rule_title_font = Font(name="微软雅黑", size=11, bold=True, color="2563eb")
            rule_title_alignment = Alignment(horizontal="left", vertical="center")

            header_font = Font(name="微软雅黑", size=10, bold=True)
            header_fill = PatternFill(start_color="e5e7eb", end_color="e5e7eb", fill_type="solid")
            header_alignment = Alignment(horizontal="left", vertical="center")
            content_font = Font(name="微软雅黑", size=10)
            content_alignment = Alignment(horizontal="left", vertical="center")
            red_font = Font(name="微软雅黑", size=10, color="FF0000")
            blue_font = Font(name="微软雅黑", size=10, color="2563eb")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            default_col_widths = {
                "部门受案号": 20,
                "案件名称": 25,
                "嫌疑人姓名": 10,
                "受理日期": 12,
                "承办部门": 15,
                "承办检察官": 12,
                "嫌疑人编号": 15,
            }

            current_row = 1

            for code in sorted(rule_group.keys()):
                rule_data = rule_group[code]
                if not rule_data:
                    continue

                rule_code = rule_data[0].get("规则编号", code)
                rule_name = rule_data[0].get("规则名称", "")
                title_cell = ws.cell(row=current_row, column=1, value=f"{rule_code} {rule_name}")
                title_cell.font = rule_title_font
                title_cell.alignment = rule_title_alignment
                # 合并第1-10列
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.row_dimensions[current_row].height = 22
                current_row += 1

                all_cols = []
                for e in rule_data:
                    for k in e.keys():
                        if k not in all_cols:
                            all_cols.append(k)

                # 按顺序排列列，确保来源表和冲突说明在最后两列
                priority_order = ["来源表", "冲突说明"]
                other_cols = [c for c in all_cols if c not in ("规则编号", "规则名称") and c not in priority_order]
                
                header_cols = other_cols + priority_order

                for col_num, col_name in enumerate(header_cols, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 22
                current_row += 1

                for row_data in rule_data:
                    for col_num, col_name in enumerate(header_cols, 1):
                        value = row_data.get(col_name, "")
                        cell = ws.cell(row=current_row, column=col_num, value=value)
                        if col_name == "冲突说明":
                            cell.font = blue_font
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            ws.row_dimensions[current_row].height = 22
                        elif col_name not in ["部门受案号", "案件名称", "嫌疑人姓名", "受理日期", "承办部门", "承办检察官", "嫌疑人编号", "来源表"]:
                            cell.font = red_font
                            cell.alignment = content_alignment
                            ws.row_dimensions[current_row].height = 22
                        else:
                            cell.font = content_font
                            cell.alignment = content_alignment
                            ws.row_dimensions[current_row].height = 22
                        cell.border = thin_border
                    current_row += 1

                current_row += 1

            for col_num, col_name in enumerate(header_cols, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = default_col_widths.get(col_name, 10)

            for col_num in range(8, 21):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 15

            for row_idx in range(1, current_row):
                ws.row_dimensions[row_idx].height = 22

            wb.save(fn)
            
            try:
                if status_vars and "result" in status_vars:
                    status_vars["result"].set(f"✅ 完成，已导出：{fn}")
                if status_vars and "conflicts" in status_vars:
                    status_vars["conflicts"].set(f"{len(all_err)} 条")
                if status_vars and "status" in status_vars:
                    status_vars["status"].set(f"✅ 校验完成！发现 {len(all_err)} 条冲突")
                if status_label:
                    status_label.config(bg="#22c55e", fg="#ffffff")
            except:
                pass
            
            try:
                os.startfile(os.path.abspath(fn))
            except:
                pass
        else:
            try:
                if status_vars and "result" in status_vars:
                    status_vars["result"].set("✅ 完成，无冲突")
                if status_vars and "conflicts" in status_vars:
                    status_vars["conflicts"].set("0 条")
                if status_vars and "status" in status_vars:
                    status_vars["status"].set("✅ 校验完成！未发现冲突")
                if status_label:
                    status_label.config(bg="#22c55e", fg="#ffffff")
            except:
                pass
    except Exception as e:
        show_custom_dialog("错误", f"校验出错: {str(e)}", "error")
        import traceback
        traceback.print_exc()


# ====================== 自定义按钮 ======================
def create_custom_button(parent, text, command, bg="#2563eb", active_bg="#1e40af", 
                          fg="white", active_fg="white", width=None, font_size=10):
    btn = tk.Button(
        parent,
        text=text,
        command=command,
        font=("微软雅黑", font_size, "bold"),
        bg=bg,
        fg=fg,
        activebackground=active_bg,
        activeforeground=active_fg,
        relief="flat",
        cursor="hand2",
        padx=16,
        pady=8
    )
    if width: btn.config(width=width)
    return btn

# ====================== 界面管理 ======================
def clear_win():
    for w in root.winfo_children(): w.destroy()

def show_renew_window():
    t = tk.Toplevel(root)
    t.title("激活续费")
    t.geometry("550x420")
    t.resizable(False, False)
    t.configure(bg="#f8fafc")
    
    t.update_idletasks()
    w = t.winfo_width()
    h = t.winfo_height()
    x = (t.winfo_screenwidth() // 2) - (w // 2)
    y = (t.winfo_screenheight() // 2) - (h // 2)
    t.geometry(f"{w}x{h}+{x}+{y}")
    
    header = tk.Frame(t, bg="#1e40af", height=80)
    header.pack(fill="x")
    header.pack_propagate(False)
    tk.Label(header, text="🔑 激活续费", font=("微软雅黑", 20, "bold"), 
             bg="#1e40af", fg="white").pack(pady=25)
    
    content = tk.Frame(t, bg="#f8fafc", padx=40, pady=30)
    content.pack(fill="both", expand=True)
    
    tk.Label(content, text="识别码", font=("微软雅黑", 12, "bold"), 
             bg="#f8fafc", fg="#374151").pack(anchor="w", pady=(0, 8))
    device_frame = tk.Frame(content, bg="#f1f5f9", bd=1, relief="solid")
    device_frame.pack(fill="x", pady=(0, 20))
    tk.Label(device_frame, text=generate_short_identifier(), font=("Consolas", 14), 
             bg="#f1f5f9", fg="#1f2937", padx=15, pady=12).pack(anchor="w")
    
    tk.Label(content, text="激活码", font=("微软雅黑", 12, "bold"), 
             bg="#f8fafc", fg="#374151").pack(anchor="w", pady=(0, 8))
    e = tk.Entry(content, font=("Consolas", 14), bg="#ffffff", fg="#1f2937", 
                 bd=1, relief="solid", insertbackground="#3b82f6")
    e.pack(fill="x", pady=(0, 25), ipady=10, ipadx=10)
    
    btn_frame = tk.Frame(content, bg="#f8fafc")
    btn_frame.pack(fill="x")
    
    def cb():
        ok, m = activate_new(e.get().strip())
        if ok:
            show_custom_dialog("激活成功", "激活成功！您可以正常使用了", "success")
            refresh_auth_label()
            t.destroy()
        else:
            show_custom_dialog("激活失败", m, "error")
    
    create_custom_button(btn_frame, "立即激活", cb, bg="#22c55e", active_bg="#16a34a", font_size=11).pack(fill="x", pady=10)

def show_active_page():
    clear_win()
    root.configure(bg="#f8fafc")
    
    main = tk.Frame(root, bg="#f8fafc")
    main.pack(fill="both", expand=True)
    
    header = tk.Frame(main, bg="#1e40af", height=130)
    header.pack(fill="x")
    header.pack_propagate(False)
    tk.Label(header, text="检察机关案卡冲突排查系统", 
             font=("微软雅黑", 22, "bold"), bg="#1e40af", fg="white").pack(pady=20)
    tk.Label(header, text="系统激活", 
             font=("微软雅黑", 14), bg="#1e40af", fg="#dbeafe").pack()
    
    content = tk.Frame(main, bg="#f8fafc")
    content.pack(fill="both", expand=True, padx=100, pady=60)
    
    card = tk.Frame(content, bg="#ffffff", bd=1, relief="ridge")
    card.pack(fill="both", expand=True)
    
    tk.Label(card, text="识别码", font=("微软雅黑", 13, "bold"), 
             bg="#ffffff", fg="#374151").pack(anchor="w", pady=(40, 10), padx=50)
    device_frame = tk.Frame(card, bg="#f1f5f9", bd=1, relief="solid")
    device_frame.pack(fill="x", padx=50, pady=(0, 30))
    tk.Label(device_frame, text=generate_short_identifier(), font=("Consolas", 16), 
             bg="#f1f5f9", fg="#1f2937", padx=20, pady=15).pack(anchor="w")
    
    tk.Label(card, text="激活码", font=("微软雅黑", 13, "bold"), 
             bg="#ffffff", fg="#374151").pack(anchor="w", pady=(0, 10), padx=50)
    e = tk.Entry(card, font=("Consolas", 14), bg="#ffffff", fg="#1f2937", 
                 bd=1, relief="solid", insertbackground="#3b82f6")
    e.pack(fill="x", padx=50, pady=(0, 40), ipady=12, ipadx=12)
    
    btn_frame = tk.Frame(card, bg="#ffffff")
    btn_frame.pack(fill="x", padx=50, pady=(0, 50))
    
    def cb():
        ok, m = activate_new(e.get().strip())
        if ok:
            show_main_page()
        else:
            show_custom_dialog("激活失败", m, "error")
    
    create_custom_button(btn_frame, "激活系统", cb, bg="#22c55e", active_bg="#16a34a", font_size=12).pack(fill="x")

def show_main_page():
    clear_win()
    global label_arrest, label_prosec, label_filter, label_auth_info, status_vars, status_label
    global label_arrest_count, label_prosec_count, label_filter_count
    global label_arrest_tip, label_prosec_tip, label_filter_tip
    auto_scan_file()
    
    bg_color = "#f0f4f8"
    root.configure(bg=bg_color)
    
    top = tk.Frame(root, bg="#1e3a8a", height=72)
    top.pack(fill="x")
    top.pack_propagate(False)
    
    title_container = tk.Frame(top, bg="#1e3a8a")
    title_container.pack(side="left", padx=30, pady=18)
    tk.Label(title_container, text="检察机关案卡冲突排查系统", 
             font=("微软雅黑", 20, "bold"), fg="#e0f2fe", bg="#1e3a8a").pack(side="left")
    
    auth_container = tk.Frame(top, bg="#1e3a8a")
    auth_container.pack(side="right", padx=30, pady=14)
    
    renew_btn = create_custom_button(auth_container, "🔑 续费", show_renew_window, 
                                      bg="#3b82f6", active_bg="#2563eb", font_size=11)
    renew_btn.pack(side="right", padx=12)
    
    label_auth_info = tk.Label(auth_container, font=("微软雅黑", 12, "bold"), 
                               padx=14, pady=8, relief="ridge", bd=1)
    label_auth_info.pack(side="right", padx=10)
    refresh_auth_label()
    
    main_paned = tk.PanedWindow(root, orient="horizontal", bg=bg_color, 
                                sashrelief="ridge", sashwidth=10, opaqueresize=True)
    main_paned.pack(fill="both", expand=True, padx=24, pady=24)
    
    left_panel = tk.Frame(main_paned, bg="#ffffff", bd=1, relief="ridge")
    main_paned.add(left_panel, width=480)
    
    rule_section = tk.Frame(left_panel, bg="#ffffff")
    rule_section.pack(fill="both", expand=True, padx=20, pady=20)
    
    tk.Label(rule_section, text="⚙️ 校验规则", font=("微软雅黑", 17, "bold"), 
             bg="#ffffff", fg="#1f2937").pack(anchor="w", pady=(0, 16))
    
    btn_container = tk.Frame(rule_section, bg="#ffffff")
    btn_container.pack(fill="x", pady=(0, 16))
    create_custom_button(btn_container, "全选", select_all_rules, 
                         bg="#3b82f6", active_bg="#2563eb", font_size=12, width=9).pack(side="left", padx=5)
    create_custom_button(btn_container, "反选", invert_select_rules, 
                         bg="#3b82f6", active_bg="#2563eb", font_size=12, width=9).pack(side="left", padx=5)
    create_custom_button(btn_container, "清空", clear_select_rules, 
                         bg="#3b82f6", active_bg="#2563eb", font_size=12, width=9).pack(side="left", padx=5)
    
    rule_container = tk.Frame(rule_section, bg="#f8fafc", bd=1, relief="solid")
    rule_container.pack(fill="both", expand=True)
    
    cv = tk.Canvas(rule_container, bg="#f8fafc", highlightthickness=0)
    scrollbar = ttk.Scrollbar(rule_container, orient="vertical", command=cv.yview)
    fr = tk.Frame(cv, bg="#f8fafc")
    fr.bind("<Configure>", lambda e: cv.configure(scrollregion=cv.bbox("all")))
    cv.create_window((0, 0), window=fr, anchor="nw")
    cv.configure(yscrollcommand=scrollbar.set)
    cv.pack(side="left", fill="both", expand=True, padx=6, pady=6)
    scrollbar.pack(side="right", fill="y")
    
    # 添加鼠标滚轮事件处理，使鼠标放在规则上时可以滚动
    def on_mousewheel(event):
        cv.yview_scroll(int(-1*(event.delta/120)), "units")
    
    cv.bind("<MouseWheel>", on_mousewheel)
    fr.bind("<MouseWheel>", on_mousewheel)
    
    rule_checks.clear()
    
    for i, (code, name, _) in enumerate(rule_list):
        v = tk.BooleanVar(value=True)
        rule_checks[code] = v
        
        rule_frame = tk.Frame(fr, bg="#f8fafc", bd=1, relief="solid")
        rule_frame.pack(fill="x", pady=4, padx=4)
        
        # 创建复选框
        cb = tk.Checkbutton(
            rule_frame,
            text=f"{code}{name}",
            variable=v,
            font=("微软雅黑", 12),
            bg="#f8fafc",
            activebackground="#d1fae5",
            activeforeground="#065f46",
            selectcolor="#ffffff",  # 多选框保持白色，不变绿色
            padx=15,
            pady=10,
            anchor="w",
            wraplength=340,
            justify="left"
        )
        cb.pack(fill="x", padx=6, pady=6, expand=False, side="left", anchor="w")
        cb.bind("<Button-1>", lambda e, i=i: on_rule_click(e, i))
        # 为规则框架和复选框添加鼠标滚轮事件处理
        rule_frame.bind("<MouseWheel>", on_mousewheel)
        cb.bind("<MouseWheel>", on_mousewheel)
        
        # 点击规则框内任何地方都能选中或取消选中
        def toggle_rule(event, checkbutton=cb, var=v):
            var.set(not var.get())
        
        rule_frame.bind("<Button-1>", toggle_rule)
        
        # 使用闭包解决变量作用域问题
        def create_update_bg(frame, checkbutton, var):
            def update_bg(*args):
                if var.get():
                    frame.config(bg="#d1fae5")
                    checkbutton.config(bg="#d1fae5")
                else:
                    frame.config(bg="#f8fafc")
                    checkbutton.config(bg="#f8fafc")
            return update_bg
        
        # 创建并绑定背景色更新函数
        update_bg = create_update_bg(rule_frame, cb, v)
        # 初始设置
        update_bg()
        # 绑定变量变化事件
        v.trace_add("write", update_bg)
    
    right_panel = tk.Frame(main_paned, bg="#ffffff", bd=1, relief="ridge")
    main_paned.add(right_panel)
    
    file_section = tk.Frame(right_panel, bg="#ffffff")
    file_section.pack(fill="x", padx=24, pady=20)
    
    tk.Label(file_section, text="📁 数据文件", font=("微软雅黑", 15, "bold"), 
             bg="#ffffff", fg="#1f2937").pack(anchor="w", pady=(0, 12))
    
    file_items = [
        ("010101表", path_arrest, upload_arrest, "📄"),
        ("020101表", path_prosec, upload_prosec, "📄"),
        ("已核对清单", path_filter, upload_filter, "📄"),
    ]
    
    for idx, (name, var, func, icon) in enumerate(file_items):
        item_frame = tk.Frame(file_section, bg="#f8fafc", bd=1, relief="solid")
        item_frame.pack(fill="x", pady=4)
        
        tk.Label(item_frame, text=icon, font=("微软雅黑", 13), 
                 bg="#f8fafc").pack(side="left", padx=12, pady=8)
        
        tk.Label(item_frame, text=name, font=("微软雅黑", 12, "bold"), 
                 bg="#f8fafc", fg="#1f2937", width=12, anchor="w").pack(side="left", padx=(0, 8))
        
        # 状态标记
        lbl = tk.Label(item_frame, text="✗", fg="#9ca3af", 
                      font=("微软雅黑", 14, "bold"), bg="#f8fafc")
        lbl.pack(side="left", padx=(4))
        
        if idx == 0:
            label_arrest = lbl
        elif idx == 1:
            label_prosec = lbl
        else:
            label_filter = lbl
        
        # 条数显示
        count_lbl = tk.Label(item_frame, text="", 
                          bg="#f8fafc", fg="#64748b",
                          font=("微软雅黑", 11),
                          width=8, anchor="w")
        count_lbl.pack(side="left", padx=(8, 0))
        
        if idx == 0:
            label_arrest_count = count_lbl
        elif idx == 1:
            label_prosec_count = count_lbl
        else:
            label_filter_count = count_lbl
        
        # 提示文字
        tip_lbl = tk.Label(item_frame, text="", 
                          bg="#f8fafc", fg="#059669",
                          font=("微软雅黑", 12))
        tip_lbl.pack(side="left", padx=(4, 0))
        
        if idx == 0:
            label_arrest_tip = tip_lbl
        elif idx == 1:
            label_prosec_tip = tip_lbl
        else:
            label_filter_tip = tip_lbl
        
        create_custom_button(item_frame, "选择", func, bg="#3b82f6", active_bg="#2563eb", font_size=11, width=9).pack(side="right", padx=12, pady=8)

    action_section = tk.Frame(right_panel, bg="#ffffff")
    action_section.pack(fill="both", expand=True, padx=24, pady=(12, 24))

    start_frame = tk.Frame(action_section, bg="#ffffff")
    start_frame.pack(fill="x", pady=(0, 16))
    create_custom_button(start_frame, "✓ 开始校验", start_check,
                        bg="#22c55e", active_bg="#16a34a", font_size=14, width=20).pack(fill="x", ipady=5)
    
    status_section = tk.Frame(action_section, bg="#0f172a", bd=2, relief="ridge")
    status_section.pack(fill="both", expand=True)
    
    header_frame = tk.Frame(status_section, bg="#1e293b")
    header_frame.pack(fill="x")
    tk.Label(header_frame, text="📊 状态信息", font=("微软雅黑", 15, "bold"), 
             bg="#1e293b", fg="#e0f2fe").pack(anchor="w", padx=20, pady=14)
    
    global status_vars
    status_vars = {}
    
    info_frame = tk.Frame(status_section, bg="#0f172a")
    info_frame.pack(fill="both", expand=True, padx=28, pady=28)
    
    for key, label_text in [
        ("info1", "📄 010101表："),
        ("info2", "📄 020101表："),
        ("info3", "📄 已核对清单："),
        ("progress", "📊 进度："),
        ("conflicts", "⚠️ 冲突："),
        ("result", "✨ 结果："),
    ]:
        line_frame = tk.Frame(info_frame, bg="#0f172a")
        line_frame.pack(fill="x", pady=3)
        
        tk.Label(line_frame, text=label_text, font=("微软雅黑", 12), 
                 bg="#0f172a", fg="#94a3b8").pack(side="left")
        
        var = tk.StringVar(value="--")
        status_vars[key] = var
        tk.Label(line_frame, textvariable=var, font=("微软雅黑", 13, "bold"), 
                 bg="#0f172a", fg="#e0f2fe").pack(side="left", padx=12)
    
    # 添加完成提示
    status_vars["status"] = tk.StringVar(value="")
    status_label = tk.Label(info_frame, textvariable=status_vars["status"], 
                           font=("微软雅黑", 12, "bold"), 
                           bg="#f59e0b", fg="#ffffff",
                           padx=24, pady=14, relief="solid", bd=1)
    status_label.pack(side="bottom", fill="x", pady=(22, 0))
    
    update_file_status()

# ====================== 启动 ======================
if __name__ == "__main__":
    def start():
        try:
            if is_license_valid():
                show_main_page()
            else:
                show_active_page()
        except:
            show_active_page()
    
    root.after(10, start)
    root.mainloop()


